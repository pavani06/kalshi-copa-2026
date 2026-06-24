"""
opcoes_model.py — Modelo de Opcoes Copa 2026
Trata cada contrato de campeao como uma opcao binaria knockout.
Computa Greeks, cenarios e sinal de saida para o book atual.
"""
import math, csv, json, os
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))

# ─── BOOK (fonte: boletagem_export-2.xlsx snapshot 15/jun) ────────────────────
BOOK = [
    {'selecao': 'Brasil',    'lado': 'V', 'contratos': 40,  'preco_medio': 12.625, 'N': 5},
    {'selecao': 'Alemanha',  'lado': 'C', 'contratos': 30,  'preco_medio': 9.00,   'N': 5},  # 21/jun: vendeu 10C @6.00 p/ Ricardo Espindola (era 40C@8.25)
    {'selecao': 'Inglaterra','lado': 'C', 'contratos': 20,  'preco_medio': 11.75,  'N': 5},
    {'selecao': 'Portugal',  'lado': 'C', 'contratos': 10,  'preco_medio': 12.0,   'N': 5},
    {'selecao': 'Marrocos',  'lado': 'C', 'contratos': 30,  'preco_medio': 2.42,   'N': 6},
    {'selecao': 'Colombia',  'lado': 'C', 'contratos': 30,  'preco_medio': 1.83,   'N': 6},
    {'selecao': 'Noruega',   'lado': 'C', 'contratos': 10,  'preco_medio': 3.0,    'N': 6},
]

# ─── CARREGAR PRECOS ──────────────────────────────────────────────────────────
def load_prices():
    kalshi, call = {}, {}
    csv_path = os.path.join(BASE, 'historico_precos.csv')
    if os.path.exists(csv_path):
        with open(csv_path, encoding='utf-8-sig') as f:
            for row in csv.reader(f):
                if row[0] == 'data_hora': continue
                nome = row[2]
                if row[3]: kalshi[nome] = float(row[3])
                if len(row) > 4 and row[4]: call[nome] = float(row[4])
    # fallback call_prices.json
    jp = os.path.join(BASE, 'call_prices.json')
    if os.path.exists(jp):
        with open(jp) as f:
            d = json.load(f)
            ticker_map = {
                'BR': 'Brasil', 'DE': 'Alemanha', 'GB': 'Inglaterra', 'PT': 'Portugal',
                'MA': 'Marrocos', 'CO': 'Colombia', 'NO': 'Noruega', 'SN': 'Senegal',
                'ES': 'Espanha', 'FR': 'Franca', 'AR': 'Argentina', 'NL': 'Holanda',
            }
            prices_raw = d.get('prices', {})
            for ticker_code, nome in ticker_map.items():
                key = f'KXMENWORLDCUP-26-{ticker_code}'
                if key in prices_raw and nome not in call:
                    call[nome] = prices_raw[key]
    return kalshi, call

# ─── MODELO POR POSICAO ───────────────────────────────────────────────────────
def compute_position(pos, kalshi, call):
    s = pos['selecao']
    p  = kalshi.get(s, 0) / 100          # prob justa (frac)
    pm = pos['preco_medio'] / 100         # preco medio compra/venda (frac)
    c_call = call.get(s)
    N  = pos['N']                         # jogos restantes ate o titulo
    C  = pos['contratos']
    sign = -1 if pos['lado'] == 'V' else 1

    if p <= 0:
        return None

    # Prob media por jogo (media geometrica)
    w = p ** (1.0 / N)

    # Pico esperado E[max] = p*(1 - ln p)   [em pp]
    peak_pp = p * (1 - math.log(p)) * 100

    # Fator de pico = pico / preco_entrada
    fator_pico = peak_pp / (pm * 100) if pm > 0 else None

    # Mark-to-Market (Kalshi justo)
    mtm_k = sign * C * (p - pm) * 100

    # Mark-to-Market (Call da Copa) — usa bid para long, ask para short
    mtm_c = sign * C * ((c_call or 0) - pm * 100) if c_call else None

    # === GREEKS ===
    # Delta de jogo: cenario se vencer o proximo jogo
    p_win_game = p / w                            # nova prob apos vencer
    delta_game_up = (p_win_game - p) * 100        # ganho pp/contrato
    scen_up = sign * C * delta_game_up            # impacto posicao

    # Delta de jogo: cenario se for eliminado
    delta_game_dn = -p * 100                      # perde toda a posicao
    scen_dn = sign * C * delta_game_dn

    # Gamma = convexidade = (1/w - 1)
    # Mede assimetria ganho/perda: quanto maior, mais knockout assimetrico
    gamma = 1.0/w - 1

    # Vega (volatilidade por jogo): DP do resultado do proximo jogo
    # SD = C * p * sqrt(1/w - 1)
    vega = C * p * math.sqrt(1.0/w - 1) * 100

    # Sinal de saida: preco atual / pico  (0%=longe; >80%=perto do pico)
    sinal_saida = (p * 100) / peak_pp

    # Edge vs Call da Copa (se disponivel)
    edge_call = (c_call - kalshi.get(s, 0)) * sign if c_call else None

    return {
        'selecao': s, 'lado': pos['lado'], 'C': C,
        'pm': pm * 100, 'N': N,
        'p_kalshi': p * 100,
        'p_call': c_call,
        'w_jogo': w * 100,
        'peak': peak_pp,
        'fator_pico': fator_pico,
        'mtm_kalshi': mtm_k,
        'mtm_call': mtm_c,
        'scen_up': scen_up,
        'scen_dn': scen_dn,
        'gamma': gamma,
        'vega': vega,
        'sinal_saida': sinal_saida,
        'edge_call': edge_call,
    }

# ─── GERAR EXCEL ─────────────────────────────────────────────────────────────
def gerar_excel(resultados, kalshi, call):
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border,
                                     Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl nao instalado")
        return

    wb = openpyxl.Workbook()

    # ── cores ──
    COR_HEADER  = "1F3864"   # azul escuro
    COR_LONG    = "D9EAD3"   # verde claro
    COR_SHORT   = "FCE5CD"   # laranja claro
    COR_TITULO  = "274E13"   # verde titulo
    COR_WARN    = "FF0000"   # vermelho exit signal
    COR_OK      = "00B050"   # verde

    def hdr(ws, row, col, val, bg=COR_HEADER, bold=True, wrap=True, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color="FFFFFF" if bg == COR_HEADER else "000000", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        return c

    def val(ws, row, col, v, fmt=None, color=None, bold=False, align="right"):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(bold=bold, color=color or "000000", size=9)
        c.alignment = Alignment(horizontal=align, vertical="center")
        if fmt: c.number_format = fmt
        return c

    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_border(ws, row_start, row_end, col_start, col_end):
        for r in range(row_start, row_end+1):
            for c in range(col_start, col_end+1):
                ws.cell(row=r, column=c).border = border

    # ── ABA 1: Opcoes Model ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Opcoes Model"
    ws1.freeze_panes = "A4"
    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 14
    ws1.row_dimensions[3].height = 30

    # Titulo
    ws1.merge_cells("A1:R1")
    t = ws1["A1"]
    t.value = f"MODELO DE OPÇÕES — COPA DO MUNDO 2026   |   {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitulo secoes
    def secao(ws, row, c1, c2, texto, bg="2D6A4F"):
        ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
        c = ws.cell(row=row, column=c1, value=texto)
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    secao(ws1, 2, 1,  5, "POSIÇÃO",        "1F3864")
    secao(ws1, 2, 6,  9, "PRECIFICAÇÃO",   "1A5276")
    secao(ws1, 2, 10, 12, "GREEKS",        "1F618D")
    secao(ws1, 2, 13, 15, "CENÁRIOS JOGO", "117A65")
    secao(ws1, 2, 16, 17, "SAÍDA",         "7D6608")
    secao(ws1, 2, 18, 18, "EDGE CALL",     "641E16")

    headers3 = [
        "País", "L/S", "Ctrl", "P.Médio", "N",        # 1-5
        "Kalshi%", "Call%", "MTM_K", "MTM_C",          # 6-9
        "w_jogo%", "Gamma", "Vega",                     # 10-12
        "+Scen", "-Scen", "Ratio",                      # 13-15
        "Peak%", "Sinal%",                              # 16-17
        "Edge_Call",                                    # 18
    ]
    for i, h in enumerate(headers3, 1):
        hdr(ws1, 3, i, h)

    port_mtm_k = 0
    port_mtm_c = 0
    port_scen_up = 0
    port_scen_dn = 0

    for r_idx, res in enumerate(resultados, 4):
        bg = COR_LONG if res['lado'] == 'C' else COR_SHORT
        fill = PatternFill("solid", fgColor=bg)
        is_short = res['lado'] == 'V'

        def dv(col, v, fmt=None, bold=False, color=None):
            c = ws1.cell(row=r_idx, column=col, value=v)
            c.fill = fill
            c.font = Font(bold=bold, size=9, color=color or "000000")
            c.alignment = Alignment(horizontal="right" if col > 2 else "left", vertical="center")
            if fmt: c.number_format = fmt
            return c

        dv(1,  res['selecao'],   bold=True, color="000000"); ws1.cell(row=r_idx, column=1).alignment = Alignment(horizontal="left")
        dv(2,  res['lado'])
        dv(3,  res['C'])
        dv(4,  res['pm'],        fmt="0.000")
        dv(5,  res['N'])
        dv(6,  res['p_kalshi'],  fmt="0.00")
        dv(7,  res['p_call'],    fmt="0.00")

        mtm_k = res['mtm_kalshi']
        mtm_c = res['mtm_call']
        mtm_k_color = "00B050" if mtm_k >= 0 else "C0392B"
        mtm_c_color = ("00B050" if mtm_c >= 0 else "C0392B") if mtm_c else "999999"
        dv(8,  mtm_k, fmt="0.0", color=mtm_k_color, bold=abs(mtm_k)>50)
        dv(9,  mtm_c, fmt="0.0", color=mtm_c_color)

        dv(10, res['w_jogo'],    fmt="0.0")
        dv(11, res['gamma'],     fmt="0.000", bold=res['gamma']>1.0)
        dv(12, res['vega'],      fmt="0.0")

        su = res['scen_up']
        sd = res['scen_dn']
        ratio = abs(su / sd) if sd != 0 else None
        su_color = "00B050" if su >= 0 else "C0392B"
        sd_color = "00B050" if sd >= 0 else "C0392B"
        dv(13, su,    fmt="+0.0;-0.0", color=su_color)
        dv(14, sd,    fmt="+0.0;-0.0", color=sd_color)
        dv(15, ratio, fmt="0.00x" if ratio else None)

        sinal = res['sinal_saida']
        sinal_color = "C0392B" if sinal > 0.8 else ("E67E22" if sinal > 0.6 else "000000")
        dv(16, res['peak'],      fmt="0.00")
        dv(17, sinal,            fmt="0%", color=sinal_color, bold=sinal > 0.7)

        edge = res['edge_call']
        edge_color = "00B050" if edge and edge > 0 else "C0392B"
        dv(18, edge,             fmt="+0.00;-0.00", color=edge_color if edge else "999999")

        port_mtm_k += mtm_k
        if mtm_c: port_mtm_c += mtm_c
        port_scen_up += su
        port_scen_dn += sd

    # Linha de total
    tr = len(resultados) + 4
    tot_fill = PatternFill("solid", fgColor="BDC3C7")
    for c in range(1, 19):
        ws1.cell(row=tr, column=c).fill = tot_fill
    for c_idx, v in [(1, "PORTFOLIO"), (8, port_mtm_k), (9, port_mtm_c or None),
                     (13, port_scen_up), (14, port_scen_dn)]:
        cell = ws1.cell(row=tr, column=c_idx, value=v)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="right" if c_idx > 1 else "left")
        if c_idx > 1 and v is not None:
            cell.number_format = "+0.0;-0.0"
            cell.font = Font(bold=True, size=9,
                             color="00B050" if v >= 0 else "C0392B")

    # Larguras
    widths = [12, 4, 4, 7, 3, 7, 7, 8, 8, 7, 7, 7, 8, 8, 6, 7, 6, 8]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    apply_border(ws1, 3, tr, 1, 18)

    # ── ABA 2: Cenarios ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Cenarios")
    ws2.freeze_panes = "B3"
    ws2.merge_cells("A1:I1")
    t2 = ws2["A1"]
    t2.value = "MATRIZ DE CENÁRIOS — Impacto no Portfolio (pp×contratos)"
    t2.font = Font(bold=True, size=11, color="FFFFFF")
    t2.fill = PatternFill("solid", fgColor="1F3864")
    t2.alignment = Alignment(horizontal="center")

    cols_scen = ["País", "Lado", "Ctrl"] + [r['selecao'] for r in resultados]
    for i, h in enumerate(cols_scen, 1):
        hdr(ws2, 2, i, h)

    # cada linha = uma posicao; cada coluna (4+) = cenario se ESSE time vencer
    nomes = [r['selecao'] for r in resultados]
    for r_idx, res in enumerate(resultados, 3):
        ws2.cell(row=r_idx, column=1, value=res['selecao']).font = Font(bold=True, size=9)
        ws2.cell(row=r_idx, column=2, value=res['lado']).font = Font(size=9)
        ws2.cell(row=r_idx, column=3, value=res['C']).font = Font(size=9)
        for c_idx, nome in enumerate(nomes, 4):
            if nome == res['selecao']:
                # this position: +scen_up
                v = res['scen_up']
            else:
                v = 0   # other team winning doesn't affect this position
            cell = ws2.cell(row=r_idx, column=c_idx, value=v if v != 0 else "-")
            cell.font = Font(size=9, color="00B050" if (isinstance(v,float) and v>0) else ("C0392B" if (isinstance(v,float) and v<0) else "999999"))
            cell.alignment = Alignment(horizontal="right")
            if isinstance(v, float): cell.number_format = "+0.0;-0.0"

    # Linhas de TOTAL e ELIMINADO
    # total se time X vencer prox jogo
    tr2 = len(resultados) + 3
    ws2.cell(row=tr2, column=1, value="SE VENCER (port.)").font = Font(bold=True, size=9)
    for c_idx, res in enumerate(resultados, 4):
        v = res['scen_up']
        cell = ws2.cell(row=tr2, column=c_idx, value=v)
        cell.font = Font(bold=True, size=9, color="00B050" if v >= 0 else "C0392B")
        cell.number_format = "+0.0;-0.0"

    tr3 = tr2 + 1
    ws2.cell(row=tr3, column=1, value="SE ELIMINAR (port.)").font = Font(bold=True, size=9)
    for c_idx, res in enumerate(resultados, 4):
        v = res['scen_dn']
        cell = ws2.cell(row=tr3, column=c_idx, value=v)
        cell.font = Font(bold=True, size=9, color="00B050" if v >= 0 else "C0392B")
        cell.number_format = "+0.0;-0.0"

    for i in range(1, len(nomes)+4):
        ws2.column_dimensions[get_column_letter(i)].width = 11

    apply_border(ws2, 2, tr3, 1, len(nomes)+3)

    # ── ABA 3: Glossario ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Glossario")
    gloss = [
        ("MÉTRICA", "FORMULA / DEFINIÇÃO", "COMO USAR"),
        ("Kalshi%", "Prob justa Kalshi de-vigada", "Valor teórico do contrato. Referência de saída."),
        ("Call%", "Preço médio no Call da Copa (bid do grupo)", "Compara com Kalshi para achar edge."),
        ("MTM_K", "sign × C × (Kalshi − P.Médio)", "P&L marcado a mercado vs Kalshi. Verde=ganho."),
        ("MTM_C", "sign × C × (Call − P.Médio)", "P&L vs preço real do Call da Copa."),
        ("N", "Jogos restantes até o título", "Ajustar a cada fase: grupos≈6, oitavas=5, QF=4, SF=3, Final=2."),
        ("w_jogo%", "p^(1/N) — prob média por jogo", "Prob estimada de vencer o PRÓXIMO jogo."),
        ("Gamma", "(1/w − 1) — convexidade", "Mede assimetria: underdog tem gamma alto (ex: Senegal γ=1.49). Long convexidade = bom."),
        ("Vega", "C × p × √(1/w − 1)", "Volatilidade da posição: DP do P&L no próximo jogo."),
        ("+Scen", "sign × C × (p/w − p)", "P&L se o time VENCER o próximo jogo."),
        ("-Scen", "sign × C × (−p)", "P&L se o time for ELIMINADO agora."),
        ("Ratio", "|+Scen| / |−Scen|", "Assimetria ganho/perda. >1 = mais a ganhar do que a perder."),
        ("Peak%", "p × (1 − ln p)", "E[max] = prob máxima esperada durante o torneio. Alvo de saída."),
        ("Sinal%", "Kalshi% / Peak%", "Distância do pico. >80% = avaliar saída agora."),
        ("Edge_Call", "Call − Kalshi (ajustado lado)", "Diferença vs mercado eficiente. Positivo = Call da Copa precifica mal a favor."),
    ]
    for r_idx, (a, b, c) in enumerate(gloss, 1):
        ws3.cell(row=r_idx, column=1, value=a).font = Font(bold=(r_idx==1), size=9)
        ws3.cell(row=r_idx, column=2, value=b).font = Font(size=9)
        ws3.cell(row=r_idx, column=3, value=c).font = Font(size=9)
        if r_idx == 1:
            for col in [1,2,3]:
                c2 = ws3.cell(row=1, column=col)
                c2.fill = PatternFill("solid", fgColor="1F3864")
                c2.font = Font(bold=True, size=9, color="FFFFFF")
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 50

    # ── SALVAR ────────────────────────────────────────────────────────────────
    out = os.path.join(BASE, "Opcoes_Copa2026.xlsx")
    wb.save(out)
    print(f"[OK] Salvo em: {out}")
    return out

# ─── PRINT CONSOLE ────────────────────────────────────────────────────────────
def print_resumo(resultados):
    print(f"\n{'='*90}")
    print(f"  MODELO DE OPÇÕES — COPA 2026   {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*90}")
    hdr = f"{'País':<12} {'L/S':>3} {'C':>3} {'P.med':>6} {'Kls%':>5} {'w%':>5} {'Peak':>5} | {'MTM_K':>7} | {'Scen+':>7} {'Scen-':>7} | {'Gma':>5} | {'Sig':>5}"
    print(hdr)
    print("-"*90)
    pt_k = pt_u = pt_d = 0
    for r in resultados:
        print(f"{r['selecao']:<12} {r['lado']:>3} {r['C']:>3} {r['pm']:>6.2f} {r['p_kalshi']:>5.2f} "
              f"{r['w_jogo']:>5.1f} {r['peak']:>5.2f} | {r['mtm_kalshi']:>+7.1f} | "
              f"{r['scen_up']:>+7.1f} {r['scen_dn']:>+7.1f} | {r['gamma']:>5.3f} | {r['sinal_saida']:>5.0%}")
        pt_k += r['mtm_kalshi']; pt_u += r['scen_up']; pt_d += r['scen_dn']
    print("-"*90)
    print(f"{'PORTFOLIO':<12} {'':>3} {'':>3} {'':>6} {'':>5} {'':>5} {'':>5} | {pt_k:>+7.1f} | {pt_u:>+7.1f} {pt_d:>+7.1f}")
    print(f"\nMTM total vs Kalshi: {pt_k:+.1f} pp×contratos")
    print(f"Se todos avancam (melhor caso): {pt_u:+.1f}")
    print(f"Se todos eliminados (pior caso): {pt_d:+.1f}")

# ─── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    kalshi, call = load_prices()
    resultados = [r for pos in BOOK if (r := compute_position(pos, kalshi, call))]
    print_resumo(resultados)
    gerar_excel(resultados, kalshi, call)


# ─── DADOS ESTRATÉGICOS (chaveamento e análise por time) ─────────────────────
CAMINHOS = {
    'Brasil': {
        'grupo': 'C', 'pts': 1, 'gd': 0, 'jogados': 1,
        'resultado_jogo1': 'Empate 1-1 vs Marrocos (13/jun)',
        'prox_jogo': 'Haiti — 19/jun (HOJE!)',
        'jogo3': 'Escócia — 24/jun',
        'posicao_esperada': '1C (se vencer Haiti)',
        'r32_como_1': 'M76 vs 2F (Holanda runner-up)',
        'r32_como_2': 'M75 vs 1F (Holanda campeão grupo — DIFÍCIL)',
        'r16_potencial': '1C: M91 vs 2E (Ivory Coast/Ecuador) ou 2I (Senegal/Noruega)',
        'qf_potencial': 'M99 vs vencedor México × Inglaterra — CONFRONTO MASSIVO',
        'risco_especial': 'COLISÃO QF COM INGLATERRA: Brasil (1C) e Inglaterra (1L) estão no MESMO LADO do chaveamento! QF provável = BR vs ENG. Natural hedge com long ENG no portfolio.',
        'dificuldade': '★★★★☆',
        'n_atualizado': 7,
        'acao': 'MANTER SHORT. Call ainda 11.75% vs Kalshi 6.29% — gap enorme. Monitorar hoje vs Haiti: se vencer e subir para ~9% Kalshi, avaliar cobertura parcial.',
    },
    'Alemanha': {
        'grupo': 'E', 'pts': 3, 'gd': 6, 'jogados': 1,
        'resultado_jogo1': 'Vitória 7-1 vs Curaçao (14/jun)',
        'prox_jogo': 'Ivory Coast — ~20/jun',
        'jogo3': 'Ecuador — 25/jun',
        'posicao_esperada': '1E (favorita absoluta no grupo)',
        'r32_como_1': 'M74 vs 3rd fraco (grupos A/B/C/D/F)',
        'r32_como_2': 'M78 vs 2I (runner-up França/Noruega)',
        'r16_potencial': '1E: M89 vs 1I — FRANÇA ou NORUEGA (brutal R16!)',
        'qf_potencial': 'M97 vs W90 (Holanda ou Brasil runner-up)',
        'risco_especial': 'Caminho duro: se Alemanha 1E e França 1I → mata-mata ALEMANHA × FRANÇA no R16. Se Noruega 1I → Alemanha vs Noruega R16. Independente, gigante no R16.',
        'dificuldade': '★★★★★',
        'n_atualizado': 7,
        'acao': 'AGUARDAR. Resultado vs Ivory Coast (próx jogo) crítico para confirmar 1E. Se Alemanha 1E e Noruega 1I → o R16 fica mais viável (Noruega < França). Edge negativo no Call (-25pp) — o grupo está pessimista vs Kalshi, correto dado o caminho duro.',
    },
    'Inglaterra': {
        'grupo': 'L', 'pts': 3, 'gd': 2, 'jogados': 1,
        'resultado_jogo1': 'Vitória 4-2 vs Croácia (jogo 1)',
        'prox_jogo': 'Ghana ou Panamá — próx. semana',
        'jogo3': 'Panana ou Ghana — jun 25-27',
        'posicao_esperada': '1L (virtual)',
        'r32_como_1': 'M80 vs 3rd fraco (E/H/I/J/K)',
        'r32_como_2': 'M83 vs 2K (runner-up Portugal/Colombia)',
        'r16_potencial': '1L: M92 vs 1A (México em casa!) — JOGO DIFÍCIL',
        'qf_potencial': 'M99 vs W91 = BRASIL (1C)! QF potencial = ENG × BRA',
        'risco_especial': 'QF INGLATERRA × BRASIL: Long ENG + Short BRA criam hedge natural. Se ENG bate BRA no QF → short BRA ganha + long ENG sobe. Cenário ideal do portfolio.',
        'dificuldade': '★★★★☆',
        'n_atualizado': 7,
        'acao': 'MANTER LONG. Kaminho duro (México R16, potencial Brasil QF) justifica preço conservador (11.97% Kalshi). Call 12.5% leve acima. Posição pequena (20 ctrl) — avaliar reforço se ENG vencer folgado o próximo jogo.',
    },
    'Portugal': {
        'grupo': 'K', 'pts': 1, 'gd': 0, 'jogados': 1,
        'resultado_jogo1': 'Empate 1-1 vs DR Congo (17/jun) — CHOQUE',
        'prox_jogo': 'Uzbequistão — 23/jun (deve vencer)',
        'jogo3': 'Colômbia — 27/jun (DECISIVO)',
        'posicao_esperada': '1K ou 2K — depende resultado vs Colombia',
        'r32_como_1': 'M87 vs 3rd de D/E/I/J/L',
        'r32_como_2': 'M83 vs 2L (Croácia — eliminada de facto)',
        'r16_potencial': '1K: M96 vs 1B (Canadá — anfitrião) | 2K: R16 vs FRANÇA (1I)',
        'qf_potencial': '1K: M100 — caminho relativamente suave | 2K: via França muito duro',
        'risco_especial': 'Portugal e Colômbia SE ENFRENTAM no jogo 3 (27/jun). Vencedor provavelmente fica em 1K (caminho suave). Perdedor fica em 2K e pode pegar França. Portfolio tem long em AMBOS — confronto direto cria risco de cancelamento.',
        'dificuldade': '★★★☆☆ (1K) / ★★★★★ (2K via França)',
        'n_atualizado': 7,
        'acao': 'ATENÇÃO: Portugal e Colombia se enfrentam na rodada final (27/jun). Long nos dois cria risco de hedge interno — um deles vai perder. Considere reduzir a posição menor (Portugal 10 ctrl) antes desse jogo se os preços subirem.',
    },
    'Marrocos': {
        'grupo': 'C', 'pts': 1, 'gd': 0, 'jogados': 1,
        'resultado_jogo1': 'Empate 1-1 vs Brasil (13/jun)',
        'prox_jogo': 'Escócia — ~22/jun (CRÍTICO — Escócia líder com 3pts)',
        'jogo3': 'Haiti — ~24/jun',
        'posicao_esperada': '2C (se vencer Escócia e Brasil ganhar Haiti)',
        'r32_como_1': 'M76 vs 2F (Holanda runner-up)',
        'r32_como_2': 'M75 vs 1F (Holanda campeão grupo — MUITO DIFÍCIL)',
        'r16_potencial': '2C: M90 = QF path vs Alemanha ou Brasil runner-up',
        'qf_potencial': 'M97 vs W89 = Alemanha ou Noruega/França area',
        'risco_especial': 'Marrocos vs Escócia (próx jogo) é praticamente eliminatório. Escócia lidera o grupo com 3pts. Derrota = Marrocos quase eliminado. Vitória = Marrocos vivo como 2C.',
        'dificuldade': '★★★★★ (grupo + R32 vs Holanda)',
        'n_atualizado': 6,
        'acao': 'POSIÇÃO DE RISCO ALTO IMEDIATO. Jogo vs Escócia é quase mata-mata. Call 3.25% vs Kalshi 1.82% — edge enorme (79%) sugere que o grupo do WhatsApp não atualizou para a dificuldade real. Se Marrocos perder vs Escócia → fechar posição imediatamente no Call.',
    },
    'Colombia': {
        'grupo': 'K', 'pts': 3, 'gd': 2, 'jogados': 1,
        'resultado_jogo1': 'Vitória 3-1 vs Uzbequistão (17/jun)',
        'prox_jogo': 'DR Congo — ~23/jun',
        'jogo3': 'Portugal — 27/jun (DECISIVO)',
        'posicao_esperada': '1K (lidera o grupo)',
        'r32_como_1': 'M87 vs 3rd de D/E/I/J/L',
        'r32_como_2': 'M83 vs 2L (Croácia)',
        'r16_potencial': '1K: M96 vs 1B (Canadá) — CAMINHO SUAVE',
        'qf_potencial': 'M100 — lado mais leve do chaveamento',
        'risco_especial': 'MELHOR CAMINHO DO PORTFOLIO se sair 1K. Evita favoritos até QF. Call 1.50% ABAIXO do Kalshi 1.72% — grupo pessimista vs mercado eficiente. Único time do portfolio com edge negativo no Call (favorável para comprar mais).',
        'dificuldade': '★★★☆☆ (1K) — o mais suave do portfolio',
        'n_atualizado': 7,
        'acao': 'MELHOR RISCO/RETORNO DO PORTFOLIO. Call (1.50%) ABAIXO do justo (1.72%) = raro edge positivo para comprar. Caminho 1K evita favoritos até QF. Considere REFORÇAR posição se Call subir acima de 2% após resultado vs DR Congo.',
    },
    'Noruega': {
        'grupo': 'I', 'pts': 3, 'gd': 3, 'jogados': 1,
        'resultado_jogo1': 'Vitória 4-1 vs Iraque (16/jun) — impressionante',
        'prox_jogo': 'Senegal — 23/jun',
        'jogo3': 'França — 26/jun (o jogo mais duro)',
        'posicao_esperada': '2I (favorita a 2º lugar, França deve ser 1I)',
        'r32_como_1': 'M77 vs 3rd fraco → R16 M89 vs ALEMANHA',
        'r32_como_2': 'M78 vs 2E (Ivory Coast/Ecuador) → R16 M89 vs FRANÇA',
        'r16_potencial': 'AMBOS os caminhos = gigante no R16 (Alemanha OU França)',
        'qf_potencial': 'M97 vs W90 = Holanda ou Brasil runner-up',
        'risco_especial': 'Noruega 1I = Alemanha R16. Noruega 2I = França R16. Erling Haaland como curinga — pode mudar a narrativa. Grupo percebe o caminho como hard e por isso Call (3.0%) vs Kalshi (2.38%) = 26% premium. Momentum é positivo.',
        'dificuldade': '★★★★★ (R16 vs Alemanha ou França garantido)',
        'n_atualizado': 7,
        'acao': 'MANTER LONG. Gamma alto (0.865). Se Noruega vencer Senegal (23/jun) → probabilidade sobe de 2.38% para ~4%+. Call vai corrigir. Saída ideal: pico durante/após o jogo vs França (26/jun) se Noruega surpreender.',
    },
    'Senegal': {
        'grupo': 'I', 'pts': 0, 'gd': -2, 'jogados': 1,
        'resultado_jogo1': 'Derrota 1-3 vs França (16/jun)',
        'prox_jogo': 'Noruega — 23/jun (PRATICAMENTE ELIMINATÓRIO)',
        'jogo3': 'Iraque — 26/jun',
        'posicao_esperada': '3I ou 4I (risco de eliminação)',
        'r32_como_1': 'Improvável dado contexto atual',
        'r32_como_2': 'Se classificar como 3º: vs winner do Grupo A/B/D/G/K/L',
        'r16_potencial': 'Improvável — precisa vencer Noruega e superar cenários',
        'qf_potencial': 'Especulativo demais no momento',
        'risco_especial': 'ALERTA DE SAÍDA: Noruega lidera com 3pts e +3 GD. Vencer Noruega é muito difícil. Derrota = Senegal eliminado com 0pts em 2 jogos. Kalshi 0.42% deve cair para 0.1% ou menos se perder.',
        'dificuldade': '★★★★★ + risco imediato de eliminação',
        'n_atualizado': 4,
        'acao': 'CONSIDERE FECHAR ANTES DE 23/JUN. Posição 10ctrl @ 0.75% vs Kalshi atual 0.42%. MTM -3.3pp. Recuperar os 4.2pp ainda disponíveis via Call ou Kalshi antes do jogo vs Noruega. Se quiser manter: aceite que a posição pode ir a zero em 4 dias.',
    },
}

def gerar_aba_caminho(wb, resultados, caminhos):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    ws = wb.create_sheet("Caminho Estratégico")
    ws.freeze_panes = "A3"

    # ── cores ──
    COR_1 = "1F3864"   # azul header
    COR_ALERTA = "C0392B"   # vermelho
    COR_OK = "1E8449"       # verde
    COR_WARN = "D68910"     # amarelo/laranja

    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titulo
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "CAMINHO ESTRATÉGICO — Copa do Mundo 2026  |  Análise de Chaveamento por Posição do Portfolio"
    t.font = Font(bold=True, size=11, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_1)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    cols = [
        "País", "L/S", "Ctrl", "Grupo", "Pts\nGD", "Situação Atual",
        "Próximo Jogo Crítico", "R32 (se 1º)", "R16 Potencial",
        "Dificuldade", "Ação Recomendada"
    ]
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COR_1)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    cor_dif = {1: "1E8449", 2: "82E0AA", 3: "F9E79F", 4: "F0A30A", 5: COR_ALERTA}

    for r_idx, res in enumerate(resultados, 3):
        s = res['selecao']
        c_data = caminhos.get(s, {})
        is_short = res['lado'] == 'V'
        bg = "FCE5CD" if is_short else "D9EAD3"
        fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[r_idx].height = 80

        def dv(col, v, bold=False, color=None, wrap=True, align="left", sz=8):
            c = ws.cell(row=r_idx, column=col, value=v)
            c.fill = fill
            c.font = Font(bold=bold, size=sz, color=color or "000000")
            c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
            c.border = border
            return c

        dv(1, s, bold=True, sz=9, align="center")
        dv(2, res['lado'], bold=True, align="center", color="C0392B" if is_short else "1E8449")
        dv(3, res['C'], align="center")
        dv(4, c_data.get('grupo', '?'), align="center", bold=True)

        pts = c_data.get('pts', '?')
        gd = c_data.get('gd', 0)
        dv(5, f"{pts}pts\nGD{gd:+d}", align="center")

        situacao = c_data.get('resultado_jogo1', '') + "\n" + c_data.get('prox_jogo', '')
        dv(6, situacao)

        prox = c_data.get('prox_jogo', '') + "\n" + c_data.get('jogo3', '')
        dv(7, prox)

        r32 = c_data.get('r32_como_1', '?')
        dv(8, r32)

        r16 = c_data.get('r16_potencial', '?')
        dv(9, r16)

        # Dificuldade com cor
        dif_str = c_data.get('dificuldade', '★★★☆☆')
        n_stars = dif_str.count('★')
        dif_color = cor_dif.get(n_stars, "000000")
        dc = ws.cell(row=r_idx, column=10, value=dif_str)
        dc.fill = PatternFill("solid", fgColor=dif_color)
        dc.font = Font(bold=True, size=9, color="FFFFFF" if n_stars >= 4 else "000000")
        dc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dc.border = border

        acao = c_data.get('acao', '')
        ac = ws.cell(row=r_idx, column=11, value=acao)
        ac.fill = fill
        ac.font = Font(size=8, bold=False, color="000000")
        ac.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ac.border = border

    # Larguras
    widths = [12, 4, 4, 6, 6, 30, 30, 28, 35, 22, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Nota de rodapé com o mapa do chaveamento
    rr = len(resultados) + 4
    ws.merge_cells(f"A{rr}:K{rr}")
    nota = ws.cell(row=rr, column=1,
        value="MAPA DO CHAVEAMENTO: R32: M73(2A×2B) M74(1E×3rd) M75(1F×2C) M76(1C×2F) M77(1I×3rd) M78(2E×2I) M79(1A×3rd) M80(1L×3rd) M83(2K×2L) M87(1K×3rd) | R16: M89=W74×W77 · M90=W73×W75 · M91=W76×W78 · M92=W79×W80 · M96=W85×W87 | QF: M97=W89×W90 · M99=W91×W92 · M100=W95×W96 | FINAL: 19/jul NY/NJ MetLife")
    nota.font = Font(size=7, italic=True, color="555555")
    nota.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[rr].height = 30


def run():
    kalshi, call = load_prices()
    resultados = [r for pos in BOOK if (r := compute_position(pos, kalshi, call))]
    print_resumo(resultados)
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(os.path.join(BASE, "Opcoes_Copa2026.xlsx"))
        # Remover aba antiga se existir
        if "Caminho Estratégico" in wb.sheetnames:
            del wb["Caminho Estratégico"]
        gerar_aba_caminho(wb, resultados, CAMINHOS)
        out = os.path.join(BASE, "Opcoes_Copa2026.xlsx")
        wb.save(out)
        print(f"[OK] Aba 'Caminho Estratégico' adicionada: {out}")
    except Exception as e:
        print(f"[ERRO] {e}")
        import traceback; traceback.print_exc()

if __name__ == "__main__":
    run()

