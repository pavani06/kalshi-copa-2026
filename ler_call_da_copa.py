#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ler_call_da_copa.py — le precos do Call da Copa colados do WhatsApp
====================================================================
Fluxo:
  1. Copie as mensagens do grupo e cole em `call_da_copa.txt`.
  2. Rode este script (ou o atualizar_call.bat).
  3. Ele reconhece a selecao + preco, grava em `call_prices.json` e
     escreve na coluna "Call da Copa %" da aba Kalshi Live do xlsx.
  4. Mostra o que casou e o que NAO foi reconhecido.

Formatos aceitos (tolerante a bagunca de chat):
  Por BANDEIRA + bid/ask (formato do grupo):
     🇧🇷 12,25 / 13      -> usa o meio (12,625)
     🇫🇷 - / 18,75       -> usa o lado disponivel (18,75)
     🏴 (Eng/Esc via tag) e 🇬🇧 tambem sao reconhecidos
  Por NOME:
     Brasil 7   Espanha 16/17   Argentina: 8   Mexico 2,5%

Precos em PONTOS PERCENTUAIS (0-100). Quando vem bid/ask, usa o MEIO.
Uso:
  python ler_call_da_copa.py                 # le call_da_copa.txt
  python ler_call_da_copa.py arquivo.txt
  python ler_call_da_copa.py --no-xlsx       # so atualiza o json
"""

import json
import os
import re
import sys
import unicodedata
import datetime as dt
from difflib import get_close_matches

HERE = os.path.dirname(os.path.abspath(__file__))
TXT = os.path.join(HERE, "call_da_copa.txt")
PRICES = os.path.join(HERE, "call_prices.json")
XLSX = os.path.join(HERE, "boletagem_export-2.xlsx")

# ticker (sufixo Kalshi) -> apelidos de texto aceitos
ALIASES = {
    "ES": ["espanha", "spain", "esp"],
    "FR": ["franca", "france", "fra"],
    "PT": ["portugal", "por"],
    "GB": ["inglaterra", "england", "eng", "ing", "uk"],
    "AR": ["argentina", "arg"],
    "BR": ["brasil", "brazil", "bra"],
    "NL": ["holanda", "paises baixos", "netherlands", "hol", "ned"],
    "DE": ["alemanha", "germany", "ale", "ger"],
    "US": ["eua", "estados unidos", "usa", "united states"],
    "NO": ["noruega", "norway", "nor"],
    "MX": ["mexico", "mex"],
    "MA": ["marrocos", "morocco", "mar"],
    "BE": ["belgica", "belgium", "bel"],
    "CO": ["colombia", "col"],
    "JP": ["japao", "japan", "jap", "jpn"],
    "IRQ": ["iraque", "iraq"],
    "COD": ["rd congo", "congo dr", "congo", "rdc"],
    "BIH": ["bosnia", "bosnia e herzegovina", "bosnia and herzegovina"],
    "CZE": ["tchequia", "republica tcheca", "czechia", "czech"],
    "CH": ["suica", "switzerland", "sui"],
    "UY": ["uruguai", "uruguay", "uru"],
    "EC": ["equador", "ecuador", "equ", "ecu"],
    "HR": ["croacia", "croatia", "cro"],
    "CIV": ["costa do marfim", "ivory coast", "marfim"],
    "SN": ["senegal", "sen"],
    "TR": ["turquia", "turkey", "tur"],
    "KR": ["coreia do sul", "south korea", "coreia", "korea"],
    "AT": ["austria", "aut"],
    "SC": ["escocia", "scotland", "esc", "sco"],
    "AU": ["australia", "aus"],
    "SE": ["suecia", "sweden", "sue", "swe"],
    "CA": ["canada", "can"],
    "DZA": ["argelia", "algeria", "alg"],
    "EGY": ["egito", "egypt", "egi", "egy"],
    "IR": ["ira", "iran"],
    "GH": ["gana", "ghana", "gha"],
    "PAN": ["panama", "pan"],
    "HTI": ["haiti", "hai"],
    "CUW": ["curacao"],
    "QAT": ["catar", "qatar", "qat"],
    "RSA": ["africa do sul", "south africa", "rsa"],
    "CPV": ["cabo verde", "cape verde"],
    "JOR": ["jordania", "jordan", "jor"],
    "UZB": ["uzbequistao", "uzbekistan", "uzb"],
    "NZL": ["nova zelandia", "new zealand", "nzl"],
    "TN": ["tunisia", "tunisia", "tun"],
    "SA": ["arabia saudita", "saudi arabia", "arabia", "ksa"],
    "PY": ["paraguai", "paraguay", "par"],
}

# ISO-2 (da bandeira) -> sufixo Kalshi, quando diferem
ISO_TO_SUF = {
    "IQ": "IRQ", "CD": "COD", "BA": "BIH", "CZ": "CZE", "PA": "PAN",
    "HT": "HTI", "CW": "CUW", "QA": "QAT", "ZA": "RSA", "CV": "CPV",
    "JO": "JOR", "UZ": "UZB", "NZ": "NZL", "DZ": "DZA", "EG": "EGY",
    "CI": "CIV", "GBENG": "GB", "GBSCT": "SC", "GBWLS": "GB",
}


def norm(s):
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9 ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


ALIAS_IDX = {}
for _tk, _names in ALIASES.items():
    for _n in _names:
        ALIAS_IDX[norm(_n)] = _tk
ALIAS_KEYS = list(ALIAS_IDX.keys())
VALID_SUF = set(ALIASES.keys())

WPP_PREFIX = re.compile(
    r"^\s*(?:\[\s*\d{1,2}[/.]\d{1,2}[/.]\d{2,4}.*?\]\s*[^:]*:\s*)"
    r"|^\s*(?:\d{1,2}[/.]\d{1,2}[/.]\d{2,4}[, ]+\d{1,2}:\d{2}(?::\d{2})?\s*[-–]\s*[^:]*:\s*)"
)
NUM = re.compile(r"\d{1,3}(?:[.,]\d+)?")
# nome + numero (+ opcional /numero) para o caminho por texto
PAIR = re.compile(
    r"([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ .'\-]{1,28}?)\s*[:\-=@]?\s*"
    r"(\d{1,3}(?:[.,]\d+)?)\s*(?:[/\- ]\s*(\d{1,3}(?:[.,]\d+)?))?\s*%?",
    re.UNICODE,
)


def decode_flag(s):
    """Le a bandeira no inicio da linha e devolve o sufixo Kalshi, ou None.
    Suporta bandeiras de indicador regional (🇧🇷) e de tag (🏴 Eng/Esc)."""
    cps = [ord(c) for c in s]
    # indicador regional (par de simbolos U+1F1E6..U+1F1FF)
    ri = [c for c in cps if 0x1F1E6 <= c <= 0x1F1FF]
    if len(ri) >= 2:
        iso = chr(ri[0] - 0x1F1E6 + 65) + chr(ri[1] - 0x1F1E6 + 65)
        return ISO_TO_SUF.get(iso, iso if iso in VALID_SUF else None)
    # sequencia de tag (bandeira preta + letras-tag): "gbeng", "gbsct"...
    if 0x1F3F4 in cps:
        tags = "".join(chr(c - 0xE0000) for c in cps if 0xE0061 <= c <= 0xE007A).upper()
        if tags:
            return ISO_TO_SUF.get(tags, tags[-2:] if tags[-2:] in VALID_SUF else None)
    return None


def mid(nums):
    """Meio do bid/ask (ou o lado disponivel). nums = lista de floats."""
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    return round(sum(nums) / len(nums), 3)


def match_team(raw):
    n = norm(raw)
    if not n:
        return None
    words = n.split()
    cands = [n] + [" ".join(words[i:]) for i in range(len(words))]
    for c in cands:
        if c in ALIAS_IDX:
            return ALIAS_IDX[c]
    for c in cands:
        m = get_close_matches(c, ALIAS_KEYS, n=1, cutoff=0.86)
        if m:
            return ALIAS_IDX[m[0]]
    return None


def parse_text(text):
    prices, matched, unmatched = {}, [], []
    for line in text.splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#"):
            continue
        clean = WPP_PREFIX.sub("", raw)

        # 1) caminho por BANDEIRA (+ bid/ask)
        suf = decode_flag(clean)
        if suf:
            nums = [float(x.replace(",", ".")) for x in NUM.findall(clean)]
            p = mid(nums)
            if p is not None and 0 <= p <= 100:
                prices[suf] = p
                matched.append((suf, "bandeira", p))
                continue
            unmatched.append(raw[:70])
            continue

        # 2) caminho por NOME
        found = False
        for m in PAIR.finditer(clean):
            team_raw, n1, n2 = m.group(1), m.group(2), m.group(3)
            tk = match_team(team_raw)
            if tk:
                nums = [float(n1.replace(",", "."))]
                if n2:
                    nums.append(float(n2.replace(",", ".")))
                p = mid(nums)
                if p is not None and 0 <= p <= 100:
                    prices[tk] = p
                    matched.append((tk, team_raw.strip(), p))
                    found = True
        if not found and re.search(r"\d", clean):
            unmatched.append(raw[:70])
    return prices, matched, unmatched


def update_xlsx(prices):
    try:
        import openpyxl
    except ImportError:
        print("[!] openpyxl nao instalado; pulei o xlsx.")
        return
    if not os.path.exists(XLSX):
        print(f"[!] Planilha nao encontrada: {XLSX}")
        return
    wb = openpyxl.load_workbook(XLSX)
    if "Kalshi Live" not in wb.sheetnames:
        print("[!] Aba 'Kalshi Live' nao existe. Rode kalshi_update.py primeiro.")
        return
    live = wb["Kalshi Live"]
    hdr = [c.value for c in live[1]]
    try:
        ci = hdr.index("Call da Copa %") + 1
        tki = hdr.index("Ticker") + 1
    except ValueError:
        print("[!] Colunas esperadas nao encontradas na aba Kalshi Live.")
        return
    n = 0
    for row in live.iter_rows(min_row=2):
        tk = row[tki - 1].value
        if not tk:
            continue
        suf = str(tk).replace("KXMENWORLDCUP-26-", "")
        if suf in prices:
            live.cell(row[0].row, ci).value = prices[suf]
            n += 1
    wb.save(XLSX)
    print(f"[ok] {n} precos escritos na coluna Call da Copa (aba Kalshi Live).")


def main():
    args = list(sys.argv[1:])
    no_xlsx = "--no-xlsx" in args
    args = [a for a in args if not a.startswith("--")]
    path = args[0] if args else TXT
    if not os.path.exists(path):
        print(f"[x] Arquivo nao encontrado: {path}")
        print("    Cole as mensagens do WhatsApp em call_da_copa.txt e rode de novo.")
        sys.exit(1)

    text = open(path, encoding="utf-8", errors="ignore").read()
    prices, matched, unmatched = parse_text(text)

    if not prices:
        print("[x] Nenhum preco reconhecido. Me mostre um trecho do formato.")
        sys.exit(1)

    prev = {}
    if os.path.exists(PRICES):
        try:
            prev = json.load(open(PRICES, encoding="utf-8")).get("prices", {})
        except Exception:
            prev = {}
    merged = dict(prev)
    merged.update({f"KXMENWORLDCUP-26-{tk}": v for tk, v in prices.items()})

    stamp = dt.datetime.now().isoformat(timespec="seconds")
    json.dump({"updated": stamp, "prices": merged}, open(PRICES, "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)

    print(f"[i] {len(matched)} precos reconhecidos:")
    for tk, raw, p in sorted(matched, key=lambda x: -x[2]):
        print(f"    {tk:4} {p:7.3f}   ({raw})")
    if unmatched:
        print(f"[!] {len(unmatched)} linha(s) com numero NAO reconhecidas:")
        for u in unmatched[:15]:
            print(f"    ? {u}")
    print(f"[ok] call_prices.json atualizado ({len(merged)} selecoes no total).")

    if not no_xlsx:
        update_xlsx({suf: v for suf, v in prices.items()})
    print("[->] Rode kalshi_update.py --no-xlsx para refletir no dashboard.")


if __name__ == "__main__":
    main()
