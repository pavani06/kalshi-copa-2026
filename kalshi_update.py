#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
kalshi_update.py — Integracao Kalshi -> Modelo do Pico (Copa 2026)
====================================================================
Puxa os dados AO VIVO do mercado "Mens World Cup Winner" (serie
KXMENWORLDCUP) da API publica da Kalshi e atualiza o projeto:

  1. Checa o status da exchange (endpoint /exchange/status).
  2. Baixa todos os ~48 mercados (1 por selecao) com bid/ask/last.
  3. Calcula a probabilidade justa (devig por normalizacao) e
     converte para odds americanas, escrevendo na coluna B do Painel
     da planilha -> todo o modelo do pico recalcula sozinho.
  4. Adiciona/atualiza a aba "Kalshi Live" com o detalhe completo.
  5. Salva um snapshot JSON e regenera o dashboard HTML.

Modos de uso:
  python kalshi_update.py                 # busca AO VIVO (precisa de internet)
  python kalshi_update.py --from-json x   # usa um dump JSON cru da API
  python kalshi_update.py --no-xlsx       # so dashboard + snapshot
  python kalshi_update.py --dry-run       # mostra o que faria, nao grava

Roda na SUA maquina (rede aberta). Dependencias: requests, openpyxl.
"""

import argparse
import datetime as dt
import json
import math
import os
import sys

BASE = "https://api.elections.kalshi.com/trade-api/v2"
SERIES = "KXMENWORLDCUP"
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "boletagem_export-2.xlsx")
SNAPSHOT = os.path.join(HERE, "kalshi_snapshot.json")
DASHBOARD = os.path.join(HERE, "kalshi_dashboard.html")
CALLPRICES = os.path.join(HERE, "call_prices.json")
HIST_CSV   = os.path.join(HERE, "historico_precos.csv")
NEWS_JSON  = os.path.join(HERE, "news_events.json")
NEWS_MAX   = 30

# Palavras-chave por ticker para detectar times em manchetes
TEAM_SEARCH_KEYWORDS = {
    "FR": ["França", "France", "Mbappé", "Mbappe", "Deschamps"],
    "ES": ["Espanha", "Spain", "España", "Yamal", "De la Fuente"],
    "GB": ["Inglaterra", "England", "Bellingham", "Southgate"],
    "AR": ["Argentina", "Messi", "Scaloni"],
    "PT": ["Portugal", "Ronaldo", "Cristiano", "Martinez"],
    "BR": ["Brasil", "Brazil", "Vinicius", "Ancelotti"],
    "DE": ["Alemanha", "Germany", "Nagelsmann"],
    "NL": ["Holanda", "Netherlands", "Koeman", "Oranje"],
    "US": ["EUA", "USA", "United States"],
    "NO": ["Noruega", "Norway", "Haaland"],
    "MA": ["Marrocos", "Morocco"],
    "CO": ["Colombia", "Colômbia", "James"],
}

# Mapeia o sufixo do ticker (KXMENWORLDCUP-26-XX) -> nome na coluna A do Painel.
# Robusto contra diferenca de grafia ingles/portugues.
TICKER_TO_PT = {
    "ES": "Espanha", "FR": "Franca", "PT": "Portugal", "GB": "Inglaterra",
    "AR": "Argentina", "BR": "Brasil", "NL": "Holanda", "DE": "Alemanha",
    "US": "EUA", "NO": "Noruega", "MX": "Mexico", "MA": "Marrocos",
    "BE": "Belgica", "CO": "Colombia", "JP": "Japao", "IRQ": "Iraque",
    "COD": "RD Congo", "BIH": "Bosnia", "CZE": "Tchequia", "CH": "Suica",
    "UY": "Uruguai", "EC": "Equador", "HR": "Croacia", "CIV": "Costa do Marfim",
    "SN": "Senegal", "TR": "Turquia", "KR": "Coreia do Sul", "AT": "Austria",
    "SC": "Escocia", "AU": "Australia", "SE": "Suecia", "CA": "Canada",
    "DZA": "Argelia", "EGY": "Egito", "IR": "Ira", "GH": "Gana",
    "PAN": "Panama", "HTI": "Haiti", "CUW": "Curacao", "QAT": "Catar",
    "RSA": "Africa do Sul", "CPV": "Cabo Verde", "JOR": "Jordania",
    "UZB": "Uzbequistao", "NZL": "Nova Zelandia", "TN": "Tunisia",
    "SA": "Arabia Saudita", "PY": "Paraguai",
}


# ----------------------------------------------------------------------------
# 1. Coleta dos dados
# ----------------------------------------------------------------------------
def fetch_exchange_status():
    import requests
    try:
        r = requests.get(f"{BASE}/exchange/status", timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": str(e)}


def fetch_live_markets():
    """Pagina /markets e devolve a lista completa de mercados da serie.
    Tenta 3 vezes; se a rede/DNS falhar, levanta ConnectionError limpo
    (sem traceback gigante) para o main() cair no ultimo snapshot."""
    import time
    import requests
    last_err = None
    for attempt in range(3):
        try:
            markets, cursor = [], None
            for _ in range(10):
                params = {"series_ticker": SERIES, "limit": 200, "status": "open"}
                if cursor:
                    params["cursor"] = cursor
                r = requests.get(f"{BASE}/markets", params=params, timeout=30)
                r.raise_for_status()
                data = r.json()
                markets.extend(data.get("markets", []))
                cursor = data.get("cursor")
                if not cursor or not data.get("markets"):
                    break
            return markets
        except requests.exceptions.RequestException as e:
            last_err = e
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
    raise ConnectionError(f"Kalshi inacessivel apos 3 tentativas (rede/DNS): {last_err}")


def load_snapshot_rows():
    """Le o ultimo snapshot salvo para uso offline. Devolve (rows, total,
    status) ou None se nao existir / estiver corrompido."""
    try:
        with open(SNAPSHOT, "r", encoding="utf-8") as f:
            d = json.load(f)
        rows = d.get("markets") or []
        if not rows:
            return None
        total = d.get("overround") or (sum(r.get("fair", 0) for r in rows) or 1.0)
        return rows, float(total), d.get("exchange_status") or {}
    except Exception as e:
        print(f"[!] Snapshot anterior indisponivel/corrompido: {e}")
        return None


def load_markets_from_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        return data.get("markets", [])
    return data


# ----------------------------------------------------------------------------
# 2. Modelo
# ----------------------------------------------------------------------------
def fair_price(m):
    """Preco justo bruto da selecao (0-1). Mid bid/ask quando ha as duas
    pontas; senao usa o last_price (longshots costumam ter bid=0)."""
    bid = float(m.get("yes_bid_dollars") or 0)
    ask = float(m.get("yes_ask_dollars") or 0)
    last = float(m.get("last_price_dollars") or 0)
    if bid > 0 and ask > 0:
        return (bid + ask) / 2.0
    if ask > 0:
        return (last + ask) / 2.0 if last > 0 else ask / 2.0
    return last


def prob_to_american(p):
    """Probabilidade (0-1) -> odds americanas. Todos os times sao
    azaroes (p<0.5) => odds positivas, igual ao input do Painel."""
    p = min(max(p, 1e-6), 0.999999)
    if p < 0.5:
        return int(round(100 * (1 - p) / p))
    return int(round(-100 * p / (1 - p)))


def build_rows(markets):
    rows = []
    for m in markets:
        tk = m.get("ticker", "")
        suf = tk.replace(f"{SERIES}-26-", "")
        fp = fair_price(m)
        if fp <= 0:
            fp = 0.001  # piso: 0,1% para nao zerar o mercado
        rows.append({
            "ticker": tk,
            "suffix": suf,
            "pt": TICKER_TO_PT.get(suf),
            "country": m.get("yes_sub_title") or suf,
            "bid": round(float(m.get("yes_bid_dollars") or 0), 4),
            "ask": round(float(m.get("yes_ask_dollars") or 0), 4),
            "last": round(float(m.get("last_price_dollars") or 0), 4),
            "fair": round(fp, 6),
            "volume": int(float(m.get("volume_fp") or 0)),
            "oi": int(float(m.get("open_interest_fp") or 0)),
        })
    total = sum(r["fair"] for r in rows) or 1.0
    for r in rows:
        r["prob_justa"] = r["fair"] / total          # devig por normalizacao
        r["odds_us"] = prob_to_american(r["fair"])    # odds cruas p/ coluna B
    rows.sort(key=lambda r: r["prob_justa"], reverse=True)
    return rows, total


def peak_metrics(p, N):
    """E[max] do martingale e teto na final (melhor caso)."""
    p = max(min(p, 0.999999), 1e-9)
    emax = p * (1 - math.log(p))
    teto = p ** (1.0 / N) if N else None
    return emax, teto


# ----------------------------------------------------------------------------
# 3. Atualizacao da planilha
# ----------------------------------------------------------------------------
def load_call_prices():
    """Le call_prices.json (precos do Call da Copa vindos do WhatsApp)."""
    try:
        d = json.load(open(CALLPRICES, encoding="utf-8"))
        return {k: v for k, v in d.get("prices", {}).items() if v not in (None, "")}
    except Exception:
        return {}


def ensure_boletagem(wb, rows):
    """Cria a aba Boletagem (log de trades + posicao + P&L) se nao existir.
    NUNCA apaga: preserva o log que o usuario ja digitou."""
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    if "Boletagem" in wb.sheetnames:
        return
    bo = wb.create_sheet("Boletagem")
    names = [r["pt"] or r["country"] for r in rows]
    LT, LB = 5, 104                 # log: linhas 5..104 (100 trades)
    PT_, PB = 5, 5 + len(names) - 1  # posicao: 48 selecoes
    hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(color="FFFFFF", bold=True)
    inblue = PatternFill("solid", fgColor="DDEBF7")

    bo["A1"] = "BOLETAGEM — Call da Copa (contratos liquidam 0-100; P&L em pontos)"
    bo["A1"].font = Font(bold=True, size=13)
    bo["A2"] = "P&L aberto (pts):"; bo["A2"].font = Font(bold=True)
    bo["C2"] = f"=SUM(N{PT_}:N{PB})"
    bo["E2"] = "Exposicao (mkt):"; bo["E2"].font = Font(bold=True)
    bo["G2"] = f"=SUM(M{PT_}:M{PB})"
    bo["I2"] = "Posicoes abertas:"; bo["I2"].font = Font(bold=True)
    bo["K2"] = f"=SUMPRODUCT(--(I{PT_}:I{PB}<>0))"

    for j, h in enumerate(["Data", "Selecao", "Lado", "Contratos", "Preco", "Obs"], 1):
        c = bo.cell(4, j, h); c.fill = hf; c.font = hfont
    for r2 in range(LT, LB + 1):
        for j in range(1, 7):
            bo.cell(r2, j).fill = inblue
        bo.cell(r2, 1).number_format = "dd/mm/yyyy"
        bo.cell(r2, 4).number_format = "0"
        bo.cell(r2, 5).number_format = "0.0"

    for j, h in enumerate(["Selecao", "Net contr.", "Custo liq.", "Preco medio",
                           "Preco atual (Call)", "Valor mkt", "P&L (pts)"]):
        c = bo.cell(4, 8 + j, h); c.fill = hf; c.font = hfont
    for idx, nm in enumerate(names):
        r2 = PT_ + idx
        bo.cell(r2, 8, nm)
        bo.cell(r2, 9).value = (f'=SUMPRODUCT(($B${LT}:$B${LB}=$H{r2})*($C${LT}:$C${LB}="Compra")*$D${LT}:$D${LB})'
                                f'-SUMPRODUCT(($B${LT}:$B${LB}=$H{r2})*($C${LT}:$C${LB}="Venda")*$D${LT}:$D${LB})')
        bo.cell(r2, 10).value = (f'=SUMPRODUCT(($B${LT}:$B${LB}=$H{r2})*($C${LT}:$C${LB}="Compra")*$D${LT}:$D${LB}*$E${LT}:$E${LB})'
                                 f'-SUMPRODUCT(($B${LT}:$B${LB}=$H{r2})*($C${LT}:$C${LB}="Venda")*$D${LT}:$D${LB}*$E${LT}:$E${LB})')
        bo.cell(r2, 11).value = f'=IF(I{r2}>0,J{r2}/I{r2},"")'
        bo.cell(r2, 12).value = f"=IFERROR(INDEX('Kalshi Live'!$I:$I,MATCH(H{r2},'Kalshi Live'!$B:$B,0)),\"\")"
        bo.cell(r2, 13).value = f'=IF(I{r2}=0,0,I{r2}*IF(L{r2}="",0,L{r2}))'
        bo.cell(r2, 14).value = f'=IF(AND(I{r2}=0,J{r2}=0),"",IF(AND(I{r2}<>0,L{r2}=""),"-",I{r2}*IF(L{r2}="",0,L{r2})-J{r2}))'
        for cc, fmt in [(11, "0.0"), (12, "0.0"), (13, "0.0"), (14, "+0.0;-0.0")]:
            bo.cell(r2, cc).number_format = fmt
    tr = PB + 1
    bo.cell(tr, 8, "TOTAL").font = Font(bold=True)
    bo.cell(tr, 13, f"=SUM(M{PT_}:M{PB})").font = Font(bold=True)
    bo.cell(tr, 14, f"=SUM(N{PT_}:N{PB})").font = Font(bold=True)
    bo.cell(tr, 13).number_format = "0.0"; bo.cell(tr, 14).number_format = "+0.0;-0.0"

    dv_lado = DataValidation(type="list", formula1='"Compra,Venda"', allow_blank=True)
    bo.add_data_validation(dv_lado); dv_lado.add(f"C{LT}:C{LB}")
    dv_sel = DataValidation(type="list", formula1=f"=$H${PT_}:$H${PB}", allow_blank=True)
    bo.add_data_validation(dv_sel); dv_sel.add(f"B{LT}:B{LB}")

    gp = PatternFill("solid", fgColor="C6EFCE"); gpf = Font(color="006100", bold=True)
    rp = PatternFill("solid", fgColor="FFC7CE"); rpf = Font(color="9C0006", bold=True)
    rng = f"N{PT_}:N{PB}"
    bo.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=gp, font=gpf))
    bo.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=rp, font=rpf))

    for col, w in {"A": 11, "B": 16, "C": 9, "D": 10, "E": 8, "F": 18, "H": 16,
                   "I": 10, "J": 11, "K": 11, "L": 16, "M": 11, "N": 11}.items():
        bo.column_dimensions[col].width = w
    bo.freeze_panes = "A5"
    bo.cell(LB + 3, 1, "Como usar: registre cada execucao no LOG (esquerda). A POSICAO (direita) soma sozinha. "
                       "Marcacao pelo preco do Call da Copa (aba Kalshi Live).")


def update_xlsx(rows, dry_run=False):
    import openpyxl
    if not os.path.exists(XLSX):
        print(f"[!] Planilha nao encontrada: {XLSX} (pulando xlsx)")
        return None
    wb = openpyxl.load_workbook(XLSX)
    call_prices = load_call_prices()
    # SEGURANCA: so escreve se a aba "Painel" existir (layout antigo, ex.
    # boletagem_export-2.xlsx). Se apontarem para um workbook sem "Painel",
    # nao escrevemos na primeira aba por engano — apenas avisamos.
    if "Painel" not in wb.sheetnames:
        print("[!] Aba 'Painel' nao existe neste workbook. Pulando escrita na "
              "planilha (use --no-xlsx). Dashboard/snapshot seguem normais.")
        return None
    ws = wb["Painel"]

    # Indexa as linhas do Painel pelo nome (coluna A), linhas 6..53
    name_to_row = {}
    for r in range(6, 60):
        v = ws.cell(r, 1).value
        if v:
            name_to_row[str(v).strip()] = r

    by_pt = {r["pt"]: r for r in rows if r["pt"]}
    updated, missing = 0, []
    for pt, r in by_pt.items():
        row = name_to_row.get(pt)
        if row:
            if not dry_run:
                ws.cell(row, 2).value = r["odds_us"]
            updated += 1
        else:
            missing.append(pt)
    unmapped = [r["country"] for r in rows if not r["pt"]]

    # Carimbo de atualizacao na linha 2
    stamp = dt.datetime.now().strftime("%d/%b/%Y %H:%M")
    if not dry_run:
        ws.cell(2, 1).value = (
            f"Odds AO VIVO da Kalshi (KXMENWORLDCUP) — atualizado {stamp}. "
            "Coluna B preenchida pelo kalshi_update.py. Ajuste N (E3) por fase e recalcule."
        )

    # Aba Kalshi Live (detalhe + metricas + edge vs Call da Copa)
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import CellIsRule
    N = ws["E3"].value or 6

    # Preserva precos do Call da Copa e o limiar entre atualizacoes
    prev_call, prev_thr = {}, 1.0
    if "Kalshi Live" in wb.sheetnames:
        old = wb["Kalshi Live"]
        hrow = [c.value for c in old[1]]
        try:
            ci = hrow.index("Call da Copa %")
            tki = hrow.index("Ticker")
            for row in old.iter_rows(min_row=2):
                tk = row[tki].value if tki < len(row) else None
                cv = row[ci].value if ci < len(row) else None
                if tk and cv not in (None, ""):
                    prev_call[str(tk)] = cv
        except ValueError:
            pass
        tcell = old["S1"].value
        if isinstance(tcell, (int, float)):
            prev_thr = tcell
        del wb["Kalshi Live"]

    live = wb.create_sheet("Kalshi Live")
    hdr = ["#", "Selecao", "Ticker", "Bid", "Ask", "Last", "Fair (bruto)",
           "Prob justa", "Call da Copa %", "Edge (pp)", "Sinal",
           "Pico esp. E[max]", "Teto final", "Mult. pico/preco",
           "Volume", "Open interest"]
    live.append(hdr)
    for i, r in enumerate(rows, 1):
        emax, teto = peak_metrics(r["prob_justa"], N)
        rr = i + 1
        live.append([
            i, r["pt"] or r["country"], r["ticker"], r["bid"], r["ask"], r["last"],
            round(r["fair"], 4), round(r["prob_justa"], 4),
            call_prices.get(r["ticker"], prev_call.get(r["ticker"])),
            f'=IF($I{rr}="","",$I{rr}-$H{rr}*100)',
            f'=IF($I{rr}="","",IF($J{rr}>$S$1,"VENDER",IF($J{rr}<-$S$1,"COMPRAR","-")))',
            round(emax, 4), round(teto, 4) if teto else None,
            round(emax / r["prob_justa"], 2) if r["prob_justa"] else None,
            r["volume"], r["oi"],
        ])
    last = len(rows) + 1

    # Parametro: limiar de edge (pp), referenciado pelas formulas de Sinal
    live["R1"] = "Limiar edge (pp):"
    live["S1"] = prev_thr
    live["R1"].font = Font(bold=True)

    # Coluna editavel Call da Copa (azul) + formatos numericos
    blue_fill = PatternFill("solid", fgColor="DDEBF7")
    blue_font = Font(color="1F4E78", bold=True)
    for rr in range(2, last + 1):
        c = live.cell(rr, 9)
        c.fill = blue_fill
        c.font = blue_font
        c.number_format = "0.0"
        live.cell(rr, 10).number_format = "+0.0;-0.0"
        live.cell(rr, 8).number_format = "0.0%"
    live["I1"].font = Font(color="1F4E78", bold=True)

    # Realce do Sinal: verde COMPRAR / vermelho VENDER
    green_f = PatternFill("solid", fgColor="C6EFCE"); green_t = Font(color="006100", bold=True)
    red_f = PatternFill("solid", fgColor="FFC7CE"); red_t = Font(color="9C0006", bold=True)
    rng = f"K2:K{last}"
    live.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"COMPRAR"'], fill=green_f, font=green_t))
    live.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"VENDER"'], fill=red_f, font=red_t))

    # Larguras
    for col, w in {"B": 16, "C": 20, "I": 14, "J": 10, "K": 11, "R": 16}.items():
        live.column_dimensions[col].width = w

    live.freeze_panes = "A2"
    note_r = last + 2
    live.cell(note_r, 1, f"Atualizado: {stamp}")
    live.cell(note_r, 3, f"N (vitorias-eq.) = {N}")
    live.cell(note_r, 5, "Edge = Call da Copa - Prob justa (pp). Digite o preco dos amigos (em %) na coluna I (azul).")

    ensure_boletagem(wb, rows)

    if not dry_run:
        import time as _time
        max_tries = 5
        for attempt in range(1, max_tries + 1):
            try:
                wb.save(XLSX)
                break
            except PermissionError:
                if attempt < max_tries:
                    print(f"[!] Planilha aberta no Excel — feche-a e aguarde... (tentativa {attempt}/{max_tries})")
                    _time.sleep(8)
                else:
                    print("[!] Nao foi possivel salvar a planilha (arquivo bloqueado).")
                    print(f"    Feche '{os.path.basename(XLSX)}' no Excel e rode o script novamente.")
                    raise
    return {"updated": updated, "missing": missing, "unmapped": unmapped, "N": N}


# ----------------------------------------------------------------------------
# 4. Historico + Noticias
# ----------------------------------------------------------------------------
def load_historico(max_teams=10):
    """Le historico_precos.csv e devolve dados prontos para Chart.js."""
    import csv
    if not os.path.exists(HIST_CSV):
        return None
    rows = []
    with open(HIST_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            try:
                rows.append({"ts": r["data_hora"], "selecao": r["selecao"],
                              "prob": float(r["kalshi_justa_pct"] or 0)})
            except Exception:
                pass
    if not rows:
        return None
    timestamps = sorted(set(r["ts"] for r in rows))
    last_ts, first_ts = timestamps[-1], timestamps[0]
    # Top times pela prob mais recente
    last_probs = {r["selecao"]: r["prob"] for r in rows if r["ts"] == last_ts}
    first_probs = {r["selecao"]: r["prob"] for r in rows if r["ts"] == first_ts}
    top_teams = sorted(last_probs, key=last_probs.get, reverse=True)[:max_teams]
    # Lookup ts -> prob por time
    team_data = {t: {} for t in top_teams}
    for r in rows:
        if r["selecao"] in top_teams:
            team_data[r["selecao"]][r["ts"]] = r["prob"]
    # Deltas: ultimo - primeiro
    deltas = {t: round(last_probs.get(t, 0) - first_probs.get(t, 0), 2) for t in top_teams}
    return {"labels": timestamps, "teams": team_data, "top_teams": top_teams, "deltas": deltas}


def detect_price_moves(min_pp=0.8):
    """Compara pontos consecutivos no historico; devolve movimentos >= min_pp."""
    import csv
    if not os.path.exists(HIST_CSV):
        return []
    rows = []
    with open(HIST_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    if len(rows) < 2:
        return []
    by_ts = {}
    for r in rows:
        ts = r["data_hora"]
        by_ts.setdefault(ts, {})[r["selecao"]] = float(r["kalshi_justa_pct"] or 0)
    timestamps = sorted(by_ts.keys())
    moves = []
    for i in range(1, len(timestamps)):
        ts_prev, ts_curr = timestamps[i - 1], timestamps[i]
        prev, curr = by_ts[ts_prev], by_ts[ts_curr]
        for team in curr:
            if team in prev:
                delta = curr[team] - prev[team]
                if abs(delta) >= min_pp:
                    try:
                        d = dt.datetime.strptime(ts_curr, "%Y-%m-%d %H:%M")
                        date_ts = d.timestamp()
                        date_str = d.strftime("%d/%m %H:%M")
                    except Exception:
                        date_ts = float(i)
                        date_str = ts_curr
                    arrow = "↑" if delta > 0 else "↓"
                    moves.append({
                        "date": date_str,
                        "date_raw": ts_curr,
                        "date_ts": date_ts,
                        "title": f"{arrow} {team}: {delta:+.1f}pp → {curr[team]:.1f}% (Kalshi)",
                        "url": "",
                        "teams": [],
                        "delta_pp": round(delta, 2),
                        "source": "price_move",
                    })
    moves.sort(key=lambda x: x["date_ts"], reverse=True)
    return moves


def fetch_news_google(queries=None):
    """Busca noticias no Google News RSS. Retorna lista de itens."""
    import requests, xml.etree.ElementTree as ET
    if queries is None:
        queries = [
            "Copa do Mundo 2026 futebol campeao",
            "World Cup 2026 winner odds prediction",
        ]
    items = []
    for q in queries:
        try:
            url = ("https://news.google.com/rss/search?q="
                   + requests.utils.quote(q)
                   + "&hl=pt-BR&gl=BR&ceid=BR:pt&num=10")
            r = requests.get(url, timeout=15,
                             headers={"User-Agent": "Mozilla/5.0 (compatible)"})
            root = ET.fromstring(r.content)
            for item in root.findall(".//item"):
                title_raw = (item.findtext("title") or "").strip()
                title = title_raw.rsplit(" - ", 1)[0].strip()
                source = (title_raw.rsplit(" - ", 1)[-1].strip()
                          if " - " in title_raw else "")
                link  = (item.findtext("link") or "").strip()
                pubdate = (item.findtext("pubDate") or "")
                try:
                    import email.utils
                    parsed = email.utils.parsedate_to_datetime(pubdate)
                    date_ts  = parsed.timestamp()
                    date_str = parsed.strftime("%d/%m %H:%M")
                except Exception:
                    date_ts, date_str = 0.0, pubdate[:16]
                # Times mencionados
                mentioned = []
                combined = title.lower()
                for tk, kws in TEAM_SEARCH_KEYWORDS.items():
                    if any(kw.lower() in combined for kw in kws):
                        mentioned.append(TICKER_TO_PT.get(tk, tk))
                items.append({
                    "date": date_str,
                    "date_ts": date_ts,
                    "title": title,
                    "source": source,
                    "url": link,
                    "teams": mentioned,
                    "source_type": "news",
                })
        except Exception as e:
            print(f"[!] Google News ({q}): {e}")
    return items


def load_news_events():
    """Le news_events.json ou retorna lista vazia."""
    try:
        with open(NEWS_JSON, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def save_and_merge_news(new_items):
    """Merge novos itens com os existentes, deduplica por titulo e salva."""
    existing = load_news_events()
    by_title = {e["title"][:80]: e for e in existing}
    for item in new_items:
        by_title[item["title"][:80]] = item
    merged = sorted(by_title.values(), key=lambda x: x.get("date_ts", 0), reverse=True)
    merged = merged[:NEWS_MAX]
    with open(NEWS_JSON, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    return merged


# ----------------------------------------------------------------------------
# 5. Snapshot + dashboard
# ----------------------------------------------------------------------------
def write_snapshot(rows, total, status, N):
    stamp = dt.datetime.now().isoformat(timespec="seconds")
    out = {
        "fetched_at": stamp,
        "series": SERIES,
        "exchange_status": status,
        "overround": round(total, 4),
        "vig_pct": round((total - 1) * 100, 2),
        "N": N,
        "markets": rows,
    }
    with open(SNAPSHOT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    return out



def render_dashboard(rows, total, status, N):
    call_prices = load_call_prices()
    stamp = dt.datetime.now().strftime("%d/%b/%Y %H:%M:%S")
    trading = (status or {}).get("trading_active")
    status_txt = ("ABERTA" if trading else "FECHADA") if isinstance(trading, bool) else "?"
    status_color = "#16a34a" if trading else "#9ca3af"
    vig = (total - 1) * 100

    hist_data  = load_historico()
    hist_json  = json.dumps(hist_data,  ensure_ascii=False) if hist_data else "null"
    news_items = load_news_events()
    news_json  = json.dumps(news_items[:25], ensure_ascii=False)

    movers_html = ""
    if hist_data and hist_data.get("deltas"):
        for team, delta in sorted(hist_data["deltas"].items(),
                                   key=lambda x: abs(x[1]), reverse=True)[:8]:
            cls  = "up" if delta >= 0 else "down"
            sign = "+" if delta >= 0 else ""
            movers_html += (f'<span class="mover-chip {cls}">'
                            f'{team} {sign}{delta:.1f}pp</span>')

    trs = []
    for i, r in enumerate(rows, 1):
        emax, teto = peak_metrics(r["prob_justa"], N)
        mult = emax / r["prob_justa"] if r["prob_justa"] else 0
        hot = "hot" if (mult >= 5 and r["prob_justa"] >= 0.01) else ""
        fair_pct = r["prob_justa"] * 100
        nm = r["pt"] or r["country"]
        cv = call_prices.get(r["ticker"])
        callval = f' value="{cv}"' if cv is not None else ""
        trs.append(f'<tr class="{hot}" data-tk="{r["suffix"]}">'
                   f'<td class="num">{i}</td><td class="nm">{nm}</td>'
                   f'<td class="num fair">{fair_pct:.1f}%</td>'
                   f'<td class="num"><input class="call" type="text" inputmode="decimal"'
                   f' data-tk="{r["suffix"]}" data-fair="{fair_pct:.4f}"{callval} placeholder="--"></td>'
                   f'<td class="num edge">--</td><td class="sig"></td>'
                   f'<td class="num">{emax*100:.1f}%</td>'
                   f'<td class="num">{teto*100:.0f}%</td>'
                   f'<td class="num mult">{mult:.1f}x</td>'
                   f'<td class="num bid">{r["bid"]*100:.1f}</td>'
                   f'<td class="num ask">{r["ask"]*100:.1f}</td></tr>')
    body = "\n".join(trs)

    # CSS parts split to avoid Edit truncation
    css_base = (
        ":root{--bg:#0b0f17;--card:#131a26;--line:#1f2a3a;--tx:#e6edf6;--mut:#8aa0bd;--ac:#4ea1ff;--hot:#f5b301}"
        "*{box-sizing:border-box}"
        "body{margin:0;background:var(--bg);color:var(--tx);font:14px/1.45 -apple-system,Segoe UI,Roboto,sans-serif;padding:24px}"
        ".wrap{max-width:980px;margin:0 auto}"
        "h1{font-size:20px;margin:0 0 4px}"
        ".sub{color:var(--mut);font-size:13px;margin-bottom:18px}"
        ".cards{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:18px}"
        ".card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:12px 16px;min-width:140px}"
        ".card .k{color:var(--mut);font-size:11px;text-transform:uppercase;letter-spacing:.04em}"
        ".card .v{font-size:20px;font-weight:600;margin-top:3px}"
        f".dot{{display:inline-block;width:9px;height:9px;border-radius:50%;background:{status_color};margin-right:6px}}"
        "table{width:100%;border-collapse:collapse;background:var(--card);border:1px solid var(--line);border-radius:12px;overflow:hidden}"
        "th,td{padding:8px 10px;border-bottom:1px solid var(--line);text-align:left}"
        "th{color:var(--mut);font-size:11px;text-transform:uppercase;letter-spacing:.03em;position:sticky;top:0;background:var(--card)}"
        "td.num{text-align:right;font-variant-numeric:tabular-nums}"
        "td.nm{font-weight:600} td.mult{color:var(--ac)}"
        "td.bid{color:#7fbf7f} td.ask{color:#e08f8f}"
        "td.fair{color:#cfe3ff;font-weight:600}"
        "tr.hot td{background:rgba(245,179,1,.07)}"
        "tr.hot td.nm::after{content:' ★';color:var(--hot)}"
        "tr:hover td{background:rgba(78,161,255,.06)}"
        "input.call{width:62px;background:#0c1422;border:1px solid #2a3a52;color:#ffd66b;"
        "border-radius:6px;padding:4px 6px;text-align:right;font:inherit;font-variant-numeric:tabular-nums}"
        "input.call:focus{outline:none;border-color:var(--ac)}"
        "td.edge{font-weight:600} td.edge.pos{color:#ff6b6b} td.edge.neg{color:#46d17f}"
        ".sig{font-size:11px;font-weight:700;letter-spacing:.03em}"
        ".sig.buy{color:#46d17f} .sig.sell{color:#ff6b6b}"
        ".sig.buy::before{content:'▲ '} .sig.sell::before{content:'▼ '}"
        ".foot{color:var(--mut);font-size:12px;margin-top:14px;line-height:1.6}"
        "a{color:var(--ac)}"
        ".bar{display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin-bottom:12px;"
        "background:var(--card);border:1px solid var(--line);border-radius:12px;padding:10px 14px}"
        ".bar label{color:var(--mut);font-size:12px}"
        ".bar input{width:54px;background:#0c1422;border:1px solid #2a3a52;color:var(--tx);"
        "border-radius:6px;padding:4px 6px;text-align:right;font:inherit}"
        ".bar .op{font-size:12px} .bar .op b{color:var(--tx)}"
        ".bar button{background:#1c2940;border:1px solid #2a3a52;color:var(--tx);border-radius:6px;"
        "padding:5px 10px;font:inherit;cursor:pointer} .bar button:hover{border-color:var(--ac)}"
    )
    css_extra = (
        ".sect{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px;margin-bottom:16px}"
        ".sect-title{font-size:12px;font-weight:700;color:var(--mut);text-transform:uppercase;letter-spacing:.05em;margin:0 0 10px}"
        ".chart-wrap{position:relative;height:300px}"
        ".movers{display:flex;flex-wrap:wrap;gap:5px;margin-bottom:10px}"
        ".mover-chip{padding:3px 9px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.01em}"
        ".mover-chip.up{background:rgba(70,209,127,.15);color:#46d17f}"
        ".mover-chip.down{background:rgba(255,107,107,.15);color:#ff6b6b}"
        ".news-list{display:flex;flex-direction:column;gap:7px}"
        ".news-item{display:flex;gap:10px;padding:9px 12px;border:1px solid var(--line);border-radius:9px;background:rgba(255,255,255,.025)}"
        ".news-item.mv-up{border-left:3px solid #46d17f}"
        ".news-item.mv-dn{border-left:3px solid #ff6b6b}"
        ".news-item.ext{border-left:3px solid var(--ac)}"
        ".news-date{color:var(--mut);font-size:11px;min-width:70px;padding-top:1px;white-space:nowrap}"
        ".news-body{flex:1}"
        ".news-title{font-size:13px;line-height:1.45}"
        ".news-title a{color:var(--tx);text-decoration:none} .news-title a:hover{color:var(--ac)}"
        ".news-tags{margin-top:4px;display:flex;flex-wrap:wrap;gap:4px}"
        ".ntag{font-size:10px;padding:1px 6px;border-radius:4px;background:rgba(78,161,255,.12);color:var(--ac)}"
        ".ntag.src{background:rgba(138,160,189,.1);color:var(--mut)}"
        ".empty{color:var(--mut);font-size:13px;text-align:center;padding:20px 0}"
    )

    js_chart = r"""
const TEAM_COLORS={
  "Franca":"#4ea1ff","Espanha":"#46d17f","Inglaterra":"#e05c5c",
  "Argentina":"#79d4f7","Portugal":"#b58ae8","Brasil":"#f5c842",
  "Alemanha":"#c0c0c0","Holanda":"#ff9c45","EUA":"#f07097",
  "Noruega":"#ff6b90","Marrocos":"#7ec8a0","Colombia":"#ffd166"
};
(function(){
  var el=document.getElementById('histChart');
  if(!HIST_DATA||!HIST_DATA.labels||HIST_DATA.labels.length===0){
    el.parentElement.innerHTML='<div class="empty">Sem dados historicos. Execute kalshi_update.py + salvar_snapshot.py para gerar historico_precos.csv</div>';
    return;
  }
  var labels=HIST_DATA.labels.map(function(ts){
    var p=ts.split(' '),d=p[0].split('-');
    return d[2]+'/'+d[1]+' '+p[1];
  });
  var datasets=HIST_DATA.top_teams.map(function(team){
    var color=TEAM_COLORS[team]||'#aaa';
    return{
      label:team,
      data:HIST_DATA.labels.map(function(ts){
        var v=HIST_DATA.teams[team]&&HIST_DATA.teams[team][ts];
        return(v!==undefined&&v!==null)?v:null;
      }),
      borderColor:color,backgroundColor:'transparent',
      tension:0.3,spanGaps:true,pointRadius:3,pointHoverRadius:5,borderWidth:2
    };
  });
  new Chart(el.getContext('2d'),{
    type:'line',
    data:{labels:labels,datasets:datasets},
    options:{
      responsive:true,maintainAspectRatio:false,
      interaction:{mode:'index',intersect:false},
      scales:{
        x:{grid:{color:'rgba(255,255,255,0.04)'},
           ticks:{color:'#8aa0bd',font:{size:10},maxRotation:40,autoSkip:true,maxTicksLimit:10}},
        y:{grid:{color:'rgba(255,255,255,0.04)'},min:0,
           ticks:{color:'#8aa0bd',font:{size:10},callback:function(v){return v.toFixed(1)+'%';}}}
      },
      plugins:{
        legend:{position:'top',labels:{color:'#e6edf6',boxWidth:13,font:{size:11},padding:14}},
        tooltip:{
          backgroundColor:'#131a26',borderColor:'#1f2a3a',borderWidth:1,
          titleColor:'#e6edf6',bodyColor:'#8aa0bd',
          callbacks:{label:function(c){return ' '+c.dataset.label+': '+(c.raw!==null?c.raw.toFixed(2)+'%':'--');}}
        }
      }
    }
  });
})();"""

    js_news = r"""
(function(){
  var nl=document.getElementById('newsList');
  if(!NEWS_EVENTS||NEWS_EVENTS.length===0){
    nl.innerHTML='<div class="empty">Sem noticias. Execute kalshi_update.py para buscar via Google News RSS.</div>';
    return;
  }
  nl.innerHTML='';
  NEWS_EVENTS.forEach(function(e){
    var isMove=e.source==='price_move';
    var delta=e.delta_pp||0;
    var cls=isMove?(delta>=0?'mv-up':'mv-dn'):'ext';
    var titleHtml=e.url
      ?'<a href="'+e.url+'" target="_blank" rel="noopener">'+e.title+'</a>'
      :e.title;
    var teamTags=(e.teams||[]).map(function(t){return '<span class="ntag">'+t+'</span>';}).join('');
    var srcTag=(e.source&&!isMove)?'<span class="ntag src">'+e.source+'</span>':'';
    nl.innerHTML+='<div class="news-item '+cls+'"><div class="news-date">'+e.date+'</div>'
      +'<div class="news-body"><div class="news-title">'+titleHtml+'</div>'
      +'<div class="news-tags">'+teamTags+srcTag+'</div></div></div>';
  });
})();"""

    js_edge = (
        "const LS_VAL='callDaCopa_v1',LS_THR='callDaCopa_thr',LS_REV='callDaCopa_rev';\n"
        f"const RENDER_REV='{stamp}';\n"
        "const $=s=>document.querySelector(s),all=s=>[...document.querySelectorAll(s)];\n"
        "function fmt(x){return (x>0?'+':'')+x.toFixed(1);}\n"
        "function recompute(){\n"
        "  const thr=parseFloat(($('#thr').value||'0').replace(',','.'))||0;\n"
        "  let store={},opps=[];\n"
        "  all('input.call').forEach(inp=>{\n"
        "    const fair=parseFloat(inp.dataset.fair);\n"
        "    const row=inp.closest('tr');\n"
        "    const ec=row.querySelector('.edge'),sc=row.querySelector('.sig');\n"
        "    const raw=(inp.value||'').trim();\n"
        "    if(raw===''){ec.textContent='--';ec.className='num edge';sc.textContent='';sc.className='sig';return;}\n"
        "    const call=parseFloat(raw.replace(',','.'));\n"
        "    if(isNaN(call)){ec.textContent='--';ec.className='num edge';sc.textContent='';sc.className='sig';return;}\n"
        "    store[inp.dataset.tk]=raw;\n"
        "    const edge=call-fair;\n"
        "    ec.textContent=fmt(edge);\n"
        "    if(edge>thr){ec.className='num edge pos';sc.textContent='VENDER';sc.className='sig sell';\n"
        "      opps.push({nm:row.querySelector('.nm').textContent,edge,kind:'sell'});}\n"
        "    else if(edge<-thr){ec.className='num edge neg';sc.textContent='COMPRAR';sc.className='sig buy';\n"
        "      opps.push({nm:row.querySelector('.nm').textContent,edge,kind:'buy'});}\n"
        "    else{ec.className='num edge';sc.textContent='--';sc.className='sig';}\n"
        "  });\n"
        "  try{localStorage.setItem(LS_VAL,JSON.stringify(store));localStorage.setItem(LS_THR,$('#thr').value);}catch(e){}\n"
        "  const buys=opps.filter(o=>o.kind==='buy').sort((a,b)=>a.edge-b.edge);\n"
        "  const sells=opps.filter(o=>o.kind==='sell').sort((a,b)=>b.edge-a.edge);\n"
        "  let parts=[];\n"
        "  if(buys.length) parts.push('<b style=\"color:#46d17f\">Comprar:</b> '+buys.slice(0,3).map(o=>o.nm+' ('+fmt(o.edge)+')').join(', '));\n"
        "  if(sells.length) parts.push('<b style=\"color:#ff6b6b\">Vender:</b> '+sells.slice(0,3).map(o=>o.nm+' ('+fmt(o.edge)+')').join(', '));\n"
        "  $('#op').innerHTML=parts.length?parts.join(' &nbsp;·&nbsp; '):'Digite os precos do Call da Copa (em %) na coluna amarela.';\n"
        "}\n"
        "function restore(){\n"
        "  try{\n"
        "    const t=localStorage.getItem(LS_THR);if(t) $('#thr').value=t;\n"
        "    if(localStorage.getItem(LS_REV)===RENDER_REV){\n"
        "      let s={};try{s=JSON.parse(localStorage.getItem(LS_VAL)||'{}');}catch(e){}\n"
        "      all('input.call').forEach(inp=>{if(s[inp.dataset.tk]!=null) inp.value=s[inp.dataset.tk];});\n"
        "    }else{\n"
        "      localStorage.setItem(LS_REV,RENDER_REV);\n"
        "      let s={};try{s=JSON.parse(localStorage.getItem(LS_VAL)||'{}');}catch(e){}\n"
        "      all('input.call').forEach(inp=>{\n"
        "        const tk=inp.dataset.tk;\n"
        "        if(inp.value!==''){s[tk]=inp.value;}\n"
        "        else if(s[tk]!=null){inp.value=s[tk];}\n"
        "      });\n"
        "      localStorage.setItem(LS_VAL,JSON.stringify(s));\n"
        "    }\n"
        "  }catch(e){}\n"
        "  recompute();\n"
        "}\n"
        "all('input.call').forEach(inp=>inp.addEventListener('input',recompute));\n"
        "$('#thr').addEventListener('input',recompute);\n"
        "$('#clr').addEventListener('click',()=>{all('input.call').forEach(i=>i.value='');try{localStorage.removeItem(LS_VAL);}catch(e){}recompute();});\n"
        "restore();\n"
    )

    html_parts = [
        "<!DOCTYPE html>",
        '<html lang="pt-BR"><head><meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width, initial-scale=1">',
        "<title>Kalshi ao vivo - Copa 2026 (Pico de probabilidade)</title>",
        '<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>',
        "<style>", css_base, css_extra, "</style></head><body><div class=\"wrap\">",
        "<h1>&#9917; Kalshi ao vivo &mdash; Mens World Cup Winner 2026</h1>",
        '<div class="sub">Modelo do pico de probabilidade de campeao &middot; serie KXMENWORLDCUP &middot; com edge vs Call da Copa</div>',
        '<div class="cards">',
        f'<div class="card"><div class="k">Exchange</div><div class="v"><span class="dot"></span>{status_txt}</div></div>',
        f'<div class="card"><div class="k">Selecoes</div><div class="v">{len(rows)}</div></div>',
        f'<div class="card"><div class="k">Overround (vig)</div><div class="v">{vig:.1f}%</div></div>',
        f'<div class="card"><div class="k">N (vitorias-eq.)</div><div class="v">{N}</div></div>',
        f'<div class="card"><div class="k">Atualizado</div><div class="v" style="font-size:14px">{stamp}</div></div>',
        "</div>",
        '<div class="sect">',
        '<div class="sect-title">&#128200; Historico de probabilidades &mdash; Kalshi</div>',
        f'<div class="movers">{movers_html}</div>',
        '<div class="chart-wrap"><canvas id="histChart"></canvas></div>',
        "</div>",
        '<div class="bar">',
        '<label>Limiar de edge (pp): <input id="thr" type="text" inputmode="decimal" value="1.0"></label>',
        '<span class="op" id="op">Digite os precos do Call da Copa (em %) na coluna amarela.</span>',
        '<button id="clr">Limpar precos</button>',
        "</div>",
        "<table><thead><tr>",
        "<th>#</th><th>Selecao</th><th>Prob justa</th><th>Call da Copa %</th><th>Edge (pp)</th><th>Sinal</th>",
        "<th>Pico E[max]</th><th>Teto</th><th>Mult.</th><th>Bid</th><th>Ask</th>",
        "</tr></thead><tbody>",
        body,
        "</tbody></table>",
        '<div class="sect" style="margin-top:18px">',
        '<div class="sect-title">&#128240; Noticias &amp; Movimentos de Preco</div>',
        '<div class="news-list" id="newsList"><div class="empty">Carregando...</div></div>',
        "</div>",
        '<div class="foot">',
        "<b>Edge = preco do Call da Copa &minus; prob justa da Kalshi</b> (em pontos percentuais). ",
        '<span style="color:#46d17f">&#9650; COMPRAR</span> = amigos abaixo do justo (barato). ',
        '<span style="color:#ff6b6b">&#9660; VENDER/FADE</span> = amigos acima do justo (caro &mdash; tipico de anfitrioes/queridinhos). ',
        "Os precos que voce digita ficam salvos neste navegador e sobrevivem as atualizacoes.<br>",
        "&#9733; = sweet spot de zebra (mult. &ge;5x e prob &ge;1%). Prob justa = Kalshi sem vig. ",
        "E[max] = p&middot;(1&minus;ln p) &middot; Teto = p^(1/N). Fonte: ",
        f'<a href="{BASE}/markets?series_ticker={SERIES}&status=open" target="_blank">API publica da Kalshi</a>. ',
        "Noticias via Google News RSS &mdash; atualiza ao rodar kalshi_update.py.",
        "</div></div>",
        "<script>",
        f"const HIST_DATA={hist_json};",
        f"const NEWS_EVENTS={news_json};",
        js_chart,
        js_news,
        js_edge,
        "</script>",
        "</body></html>",
    ]
    html = "\n".join(html_parts)
    with open(DASHBOARD, "w", encoding="utf-8") as f:
        f.write(html)
        f.flush()
        os.fsync(f.fileno())


# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Atualiza o modelo do pico com dados ao vivo da Kalshi.")
    ap.add_argument("--from-json", help="Le um dump JSON cru da API em vez de buscar ao vivo.")
    ap.add_argument("--no-xlsx",   action="store_true", help="Nao mexe na planilha.")
    ap.add_argument("--dry-run",   action="store_true", help="Mostra o resultado sem gravar a planilha.")
    ap.add_argument("--skip-news", action="store_true", help="Nao busca noticias no Google News RSS.")
    args = ap.parse_args()

    if args.from_json:
        status = {"trading_active": None, "note": "modo --from-json"}
        markets = load_markets_from_json(args.from_json)
    else:
        status = fetch_exchange_status()
        if "error" in status:
            print(f"[!] /exchange/status falhou: {status['error']}")
        else:
            print(f"[i] exchange/status: {json.dumps(status, ensure_ascii=False)}")
        try:
            markets = fetch_live_markets()
        except ConnectionError as e:
            print(f"[!] {e}")
            print("[i] Modo offline: usando o ultimo snapshot da Kalshi "
                  "(Kalshi fica defasado; o Call da Copa segue atual).")
            markets = None

    offline = False
    if markets:
        rows, total = build_rows(markets)
    else:
        fb = load_snapshot_rows()
        if not fb:
            print("[x] Sem dados ao vivo e sem snapshot valido. Abortando.")
            sys.exit(1)
        rows, total, status = fb
        offline = True
    print(f"[i] {len(rows)} selecoes - overround {total:.3f} (vig {(total-1)*100:.1f}%)")
    top = rows[0]
    print(f"[i] Favorito: {top['pt'] or top['country']} {top['prob_justa']*100:.1f}% "
          f"(bid {top['bid']*100:.0f} / ask {top['ask']*100:.0f})")

    N = 6
    if not args.no_xlsx:
        res = update_xlsx(rows, dry_run=args.dry_run)
        if res:
            N = res["N"]
            print(f"[i] Painel: {res['updated']} selecoes atualizadas (N={N}).")
            if res["missing"]:
                print(f"[!] Sem linha no Painel: {res['missing']}")
            if res["unmapped"]:
                print(f"[!] Ticker sem mapa PT: {res['unmapped']}")

    if not offline:
        write_snapshot(rows, total, status, N)
        print(f"[ok] Snapshot -> {os.path.basename(SNAPSHOT)}")
    else:
        print("[i] Snapshot preservado (offline, nao sobrescrevo dados defasados).")

    if not args.dry_run:
        print("[i] Detectando movimentos de preco no historico...")
        news_all = detect_price_moves(min_pp=0.8)
        print(f"[i] {len(news_all)} movimentos detectados.")
        if not args.skip_news:
            print("[i] Buscando noticias no Google News RSS...")
            google_items = fetch_news_google()
            print(f"[i] {len(google_items)} noticias encontradas.")
            news_all.extend(google_items)
        merged = save_and_merge_news(news_all)
        print(f"[ok] Noticias -> {os.path.basename(NEWS_JSON)} ({len(merged)} itens)")

    render_dashboard(rows, total, status, N)
    print(f"[ok] Dashboard -> {os.path.basename(DASHBOARD)}")
    if not args.dry_run and not args.no_xlsx:
        print(f"[ok] Planilha atualizada -> {os.path.basename(XLSX)}")


if __name__ == "__main__":
    main()
