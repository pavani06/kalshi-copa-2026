#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
atualizar_book.py - sincroniza o BOOK e a PRECIFICACAO a partir das fontes vivas.

Nada de campo colado na mao: tudo que depende do dia (posicoes, precos do Call,
prob justa do Kalshi) e EXTRAIDO das fontes e reescrito no Copa2026_Modelo_Pico.xlsx.

Fontes de verdade:
  - boletagem_export-2.xlsx (aba Boletagem) -> posicoes (net por pais)
  - call_da_copa.txt                       -> Call Bid/Ask vivos
  - historico_precos.csv                   -> Kalshi prob justa (%)
  - call_prices.json                       -> mid de fallback

Abas reescritas:
  - Posicoes            (auto da boletagem; ja era dinamica)
  - Plano LS 15k        (col "Ja tem R$" + "Preco ref")
  - Modelo de Pico      (inputs B/C = Call Bid/Ask, E = Kalshi Fair)
  - Comprar Convexidade (Bid/Ask, Prob; Spread/Pico/Ladder viram formula)
  - Vender Favoritos    (Bid recebe = Call bid, Prob casas = Kalshi fair)

As abas de TEXTO (Leitura Tatica, Resumo & Estrategia) NAO sao tocadas - sao
analise manual sua.

Fluxo diario:
  1. Atualize a aba Boletagem (boletagem_export-2.xlsx).
  2. Atualize os precos (atualizar_call.bat / atualizar_kalshi.bat).
  3. Rode:  python atualizar_book.py   (ou atualizar_book.bat)
"""
import os, re, csv, json, math, sys
import datetime as dt
import openpyxl
from ler_call_da_copa import decode_flag, match_team, NUM, WPP_PREFIX

HERE = os.path.dirname(os.path.abspath(__file__))
BOLETAGEM = os.path.join(HERE, "boletagem_export-2.xlsx")
MODELO    = os.path.join(HERE, "Copa2026_Modelo_Pico.xlsx")
CALL_TXT  = os.path.join(HERE, "call_da_copa.txt")
CALL_JSON = os.path.join(HERE, "call_prices.json")
HIST_CSV  = os.path.join(HERE, "historico_precos.csv")

# Expoentes do "ladder" de convexidade (preco condicional a chegar na fase).
# Ajustados aos valores originais da planilha: preco_fase% = 100 * (p/100)^exp.
PICO_EXP = {"QF": 0.663, "SEMI": 0.465, "FINAL": 0.231}

SUF_NOME = {
    "ES":"Espanha","FR":"Franca","PT":"Portugal","GB":"Inglaterra","AR":"Argentina",
    "BR":"Brasil","NL":"Holanda","DE":"Alemanha","US":"EUA","NO":"Noruega","MX":"Mexico",
    "MA":"Marrocos","BE":"Belgica","CO":"Colombia","JP":"Japao","CH":"Suica","UY":"Uruguai",
    "EC":"Equador","HR":"Croacia","CIV":"Costa do Marfim","SN":"Senegal","TR":"Turquia",
    "KR":"Coreia do Sul","AT":"Austria","SC":"Escocia","SE":"Suecia","CA":"Canada","EGY":"Egito",
}
SUF_EMOJI = {
    "ES":"\U0001F1EA\U0001F1F8","FR":"\U0001F1EB\U0001F1F7","PT":"\U0001F1F5\U0001F1F9",
    "GB":"\U0001F3F4","AR":"\U0001F1E6\U0001F1F7","BR":"\U0001F1E7\U0001F1F7",
    "NL":"\U0001F1F3\U0001F1F1","DE":"\U0001F1E9\U0001F1EA","US":"\U0001F1FA\U0001F1F8",
    "NO":"\U0001F1F3\U0001F1F4","MX":"\U0001F1F2\U0001F1FD","MA":"\U0001F1F2\U0001F1E6",
    "BE":"\U0001F1E7\U0001F1EA","CO":"\U0001F1E8\U0001F1F4","JP":"\U0001F1EF\U0001F1F5",
    "CH":"\U0001F1E8\U0001F1ED","UY":"\U0001F1FA\U0001F1FE","EC":"\U0001F1EA\U0001F1E8",
    "HR":"\U0001F1ED\U0001F1F7","SN":"\U0001F1F8\U0001F1F3",
}
TESE = {
    "DE":"Favorito mid, prob caiu. Fora da tese de convexidade - avaliar reduzir/zerar.",
    "GB":"Long OK perto do justo; NAO shortar. Vender em camadas no run-up.",
    "PT":"~no justo (Kalshi). Manter; montar escada de venda.",
    "MA":"Convexidade pura (SEMI 2022). Segurar para o pico.",
    "NO":"Convexidade (Haaland). ON-tese. Segurar.",
    "CO":"Convexidade. ON-tese. Segurar.",
    "SN":"Posicao minima; liquidez baixa - confirmar status.",
    "BR":"SHORT do favorito local (vies de torcida infla o Call). ON-tese. Risco = pico.",
}

def nome_para_suf(nome):
    suf = decode_flag(nome)
    return suf if suf else match_team(nome)

def ler_book(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if "bolet" in n.lower()), None)
    if ws is None:
        raise SystemExit("[!] Aba 'Boletagem' nao encontrada em " + path)
    hdr_row = None
    for r in range(1, 8):
        if str(ws.cell(r, 1).value).strip().lower() == "data":
            hdr_row = r; break
    if hdr_row is None:
        hdr_row = 4
    pos = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        pais = ws.cell(r, 2).value
        lado = ws.cell(r, 3).value
        ct   = ws.cell(r, 4).value
        pr   = ws.cell(r, 5).value
        if not pais or ct in (None, 0) or pr is None:
            continue
        suf = nome_para_suf(str(pais))
        if not suf:
            print("[!] linha %d: pais nao reconhecido: %r - ignorado" % (r, pais))
            continue
        sgn = 1 if str(lado).strip().lower().startswith("c") else -1
        d = pos.setdefault(suf, {"net": 0.0, "custo": 0.0})
        d["net"]   += sgn * float(ct)
        d["custo"] += sgn * float(ct) * float(pr)
    book = []
    for suf, d in pos.items():
        net = round(d["net"], 6)
        if abs(net) < 1e-9:
            continue
        book.append({
            "suf": suf, "lado": "C" if net > 0 else "V", "lot": abs(net),
            "preco_medio": round(abs(d["custo"] / net), 4),
            "custo_liq": round(d["custo"], 2),
        })
    book.sort(key=lambda x: x["lot"] * x["preco_medio"], reverse=True)
    return book

def ler_call_bidask(path):
    out = {}
    if not os.path.exists(path):
        return out
    for line in open(path, encoding="utf-8").read().splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#"):
            continue
        clean = WPP_PREFIX.sub("", raw)
        suf = decode_flag(clean)
        if not suf:
            m = re.match(r"([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ .'\-]{1,28}?)\s*[:\-=@]?\s*\d", clean)
            if m:
                suf = match_team(m.group(1))
        if not suf:
            continue
        nums = [float(x.replace(",", ".")) for x in NUM.findall(clean)]
        nums = [n for n in nums if 0 <= n <= 100]
        if not nums:
            continue
        out[suf] = (nums[0], nums[1] if len(nums) > 1 else nums[0])
    return out

def ler_call_mid(path):
    out = {}
    if os.path.exists(path):
        d = json.load(open(path, encoding="utf-8"))
        for tk, v in d.get("prices", {}).items():
            out[tk.replace("KXMENWORLDCUP-26-", "")] = float(v)
    return out

def ler_kalshi_fair(path):
    out = {}
    if not os.path.exists(path):
        return out
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            try:
                out[(row.get("ticker") or "").strip()] = float(row["kalshi_justa_pct"])
            except (TypeError, ValueError, KeyError):
                pass
    return out

def pico_esperado(fair_pct):
    if not fair_pct or fair_pct <= 0:
        return None
    return round(fair_pct * (1 - math.log(fair_pct / 100.0)), 2)

def marca(suf, callba, callmid, fallback):
    if suf in callba:
        return callba[suf]
    if callmid.get(suf):
        return callmid[suf], callmid[suf]
    return fallback, fallback

def _stamp(prefixo):
    return prefixo + dt.datetime.now().strftime("%d/%b/%Y %H:%M")

def escrever_posicoes(wb, book, callba, callmid, fair):
    ws = wb["Posicoes"]
    for r in range(4, 41):
        for c in range(1, 13):
            ws.cell(r, c).value = None
    r = 4
    for p in book:
        suf = p["suf"]
        bid, ask = marca(suf, callba, callmid, p["preco_medio"])
        pico = pico_esperado(fair.get(suf))
        ws.cell(r, 1, (SUF_EMOJI.get(suf, "") + " " + SUF_NOME.get(suf, suf)).strip())
        ws.cell(r, 2, p["lado"])
        ws.cell(r, 3, p["lot"])
        ws.cell(r, 4, p["preco_medio"])
        ws.cell(r, 5, '=IF(B%d="V",-C%d*D%d,C%d*D%d)' % (r, r, r, r, r))
        ws.cell(r, 6, round(bid, 3))
        ws.cell(r, 7, round(ask, 3))
        ws.cell(r, 8, '=IF(B%d="V",C%d*(D%d-G%d),C%d*(F%d-D%d))' % (r, r, r, r, r, r, r))
        ws.cell(r, 9, '=IFERROR(H%d/(C%d*D%d),0)' % (r, r, r))
        if pico is not None:
            ws.cell(r, 10, pico)
            ws.cell(r, 11, '=IF(B%d="V",C%d*(D%d-J%d),C%d*(J%d-D%d))' % (r, r, r, r, r, r, r))
        ws.cell(r, 12, TESE.get(suf, ""))
        ws.cell(r, 9).number_format = "0.0%"
        r += 1
    last = r - 1
    ws.cell(r, 1, "TOTAL")
    ws.cell(r, 5, "=SUM(E4:E%d)" % last)
    ws.cell(r, 8, "=SUM(H4:H%d)" % last)
    ws.cell(r, 11, "=SUM(K4:K%d)" % last)
    ws.cell(1, 1, _stamp("Posicoes (auto da boletagem) - atualizado "))
    return last

def atualizar_plano(wb, book, callba):
    """Preenche 'Ja tem R$' (do book) e 'Preco ref' (Call bid vivo) no Plano."""
    if "Plano LS 15k" not in wb.sheetnames:
        return
    ws = wb["Plano LS 15k"]
    inv = {p["suf"]: p["custo_liq"] for p in book}
    hdr_row = None; col_pais = 2; col_jatem = 3; col_ref = None
    for r in range(1, 8):
        if any("Ja tem" in str(ws.cell(r, c).value or "") for c in range(1, 9)):
            hdr_row = r
            for c in range(1, 9):
                v = str(ws.cell(r, c).value or "")
                if "Selecao" in v: col_pais = c
                if "Ja tem" in v: col_jatem = c
                if "ref" in v.lower() and "preco" in v.lower(): col_ref = c
            break
    if hdr_row is None:
        return
    for r in range(hdr_row + 1, ws.max_row + 1):
        nome = ws.cell(r, col_pais).value
        if not nome:
            continue
        suf = nome_para_suf(str(nome))
        if not suf:
            continue
        if suf in inv:
            ws.cell(r, col_jatem, round(inv[suf], 2))
        if col_ref and suf in callba:
            ws.cell(r, col_ref, round(callba[suf][0], 3))

def _itera_linhas(ws, r0, r1):
    """Gera (linha, suf) so para linhas de dados com bandeira/nome reconhecido."""
    for r in range(r0, r1 + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        suf = nome_para_suf(str(v))
        if suf:
            yield r, suf

def atualizar_modelo_pico(wb, callba, fair):
    """Modelo de Pico: B/C = Call Bid/Ask vivos, E = Kalshi Fair. Formulas intactas."""
    if "Modelo de Pico" not in wb.sheetnames:
        return 0
    ws = wb["Modelo de Pico"]
    n = 0
    for r, suf in _itera_linhas(ws, 2, 23):
        if suf in callba:
            ws.cell(r, 2, round(callba[suf][0], 3))
            ws.cell(r, 3, round(callba[suf][1], 3))
        if suf in fair:
            ws.cell(r, 5, round(fair[suf], 2))
        n += 1
    ws.cell(24, 1, _stamp("Inputs B/C (call_da_copa.txt) e E (historico_precos.csv) auto - "))
    return n

def atualizar_convexidade(wb, callba, fair):
    """Comprar Convexidade: Bid/Ask e Prob vivos; Spread/Pico/Ladder viram formula."""
    if "Comprar Convexidade" not in wb.sheetnames:
        return 0
    ws = wb["Comprar Convexidade"]
    qf, se, fi = PICO_EXP["QF"], PICO_EXP["SEMI"], PICO_EXP["FINAL"]
    n = 0
    for r, suf in _itera_linhas(ws, 4, 18):
        if suf in callba:
            ws.cell(r, 2, round(callba[suf][0], 3))   # Bid
            ws.cell(r, 3, round(callba[suf][1], 3))   # Ask
        if suf in fair:
            ws.cell(r, 5, round(fair[suf], 2))        # Prob %
        ws.cell(r, 4, f'=IF(C{r}=0,"",ROUND((C{r}-B{r})/C{r}*100,0))')          # Spread %
        ws.cell(r, 6, f'=IF(E{r}<=0,0,ROUND(E{r}*(1-LN(E{r}/100)),1))')         # Pico % (E[max])
        ws.cell(r, 8, f'=IF(E{r}<=0,"",ROUND(100*(E{r}/100)^{qf},1))')          # se QF
        ws.cell(r, 9, f'=IF(E{r}<=0,"",ROUND(100*(E{r}/100)^{se},1))')          # se SEMI
        ws.cell(r,10, f'=IF(E{r}<=0,"",ROUND(100*(E{r}/100)^{fi},1))')          # se FINAL
        n += 1
    ws.cell(19, 1, _stamp("Bid/Ask e Prob auto; Spread/Pico/ladder = formula - "))
    return n

def atualizar_vender(wb, callba, fair):
    """Vender Favoritos: Bid recebe = Call bid; Prob casas = Kalshi fair (de-vigado vivo)."""
    if "Vender Favoritos (short)" not in wb.sheetnames:
        return 0
    ws = wb["Vender Favoritos (short)"]
    n = 0
    for r, suf in _itera_linhas(ws, 4, 11):
        if suf in callba:
            ws.cell(r, 2, round(callba[suf][0], 2))   # Bid (recebe)
        if suf in fair:
            ws.cell(r, 3, round(fair[suf], 2))        # Prob casas %
        n += 1
    ws.cell(12, 1, _stamp("Bid (call) e Prob casas (Kalshi) auto; prob modelo = manual - "))
    return n

def main():
    boletagem = BOLETAGEM
    if "--boletagem" in sys.argv:
        boletagem = sys.argv[sys.argv.index("--boletagem") + 1]
    print("Lendo book da boletagem:", os.path.basename(boletagem))
    book    = ler_book(boletagem)
    callba  = ler_call_bidask(CALL_TXT)
    callmid = ler_call_mid(CALL_JSON)
    fair    = ler_kalshi_fair(HIST_CSV)
    print("\n%-16s%5s%6s%8s%7s%7s%7s%7s" % ("Pais","Lado","Lote","PrMed","Bid","Ask","Fair%","Pico%"))
    for p in book:
        suf = p["suf"]
        bid, ask = marca(suf, callba, callmid, p["preco_medio"])
        print("%-16s%5s%6.0f%8.3f%7.2f%7.2f%7.2f%7.2f" % (
            SUF_NOME.get(suf, suf), p["lado"], p["lot"], p["preco_medio"],
            bid, ask, fair.get(suf, 0), pico_esperado(fair.get(suf)) or 0))
    wb = openpyxl.load_workbook(MODELO)
    last = escrever_posicoes(wb, book, callba, callmid, fair)
    atualizar_plano(wb, book, callba)
    nmp = atualizar_modelo_pico(wb, callba, fair)
    ncv = atualizar_convexidade(wb, callba, fair)
    nvf = atualizar_vender(wb, callba, fair)
    wb.save(MODELO)
    print("\nOK -> %s reescrito:" % os.path.basename(MODELO))
    print("  Posicoes: %d posicoes | Plano LS 15k: Ja tem R$ + Preco ref" % len(book))
    print("  Modelo de Pico: %d linhas | Comprar Convexidade: %d | Vender Favoritos: %d" % (nmp, ncv, nvf))

if __name__ == "__main__":
    main()
